from __future__ import annotations
import random, os, time, uuid
from pathlib import Path
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from openpyxl import load_workbook #type: ignore

class FakeEmailjobsGenerator:
    ''' to create fake email jobs'''
    #written by AI
    BASE_DIR = Path(__file__).resolve().parent
    PIPELINE_DIR = BASE_DIR / "personal_inbox"
    INBOX_DIR = PIPELINE_DIR / "inbox"
    PROCESSING_DIR =  PIPELINE_DIR / "processing"
    ATTACHMENTS_DIR = PIPELINE_DIR / "generator_attachments"

    for folder in [PIPELINE_DIR, INBOX_DIR, PROCESSING_DIR, ATTACHMENTS_DIR]:
        folder.mkdir(exist_ok=True)


    def __init__(self) -> None:
        self.main()


    def create_example_attachment_files(self) -> None:
        """Creates a few simple test files if they do not already exist."""
        txt_path = self.ATTACHMENTS_DIR / "job1_request.txt"
        if not txt_path.exists():
            txt_path.write_text(
                "SKU=100245\nOLD_MATERIAL=MAT-OLD-778\nNEW_MATERIAL=MAT-NEW-991\n",
                encoding="utf-8",
            )

        csv_path = self.ATTACHMENTS_DIR / "job2_request.csv"
        if not csv_path.exists():
            csv_path.write_text(
                "invoice_id,action\nINV-2026-1001,close\n",
                encoding="utf-8",
            )


    def build_email_message(self,
        *,
        from_name: str,
        from_email: str,
        to_email: str,
        subject: str,
        body: str,
        attachment_paths: list[Path] | None = None,
    ) -> EmailMessage:
        msg = EmailMessage()
        msg["From"] = f"{from_name} <{from_email}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg["Date"] = formatdate(localtime=True)
        msg["Message-ID"] = make_msgid()
        msg.set_content(body)

        for path in attachment_paths or []:
            data = path.read_bytes()
            # Simple generic attachment type is enough for testing
            msg.add_attachment(
                data,
                maintype="application",
                subtype="octet-stream",
                filename=path.name,
            )

        return msg


    def write_eml_to_inbox(self,msg: EmailMessage, prefix: str = "mail") -> Path:
        """Atomic write into inbox to reduce risk of partial reads."""
        unique_id = uuid.uuid4().hex[:12]
        final_path = self.INBOX_DIR / f"{prefix}_{unique_id}.eml"
        temp_path = self.INBOX_DIR / f".tmp_{prefix}_{unique_id}.eml"

        with open(temp_path, "wb") as f:
            f.write(msg.as_bytes())

        temp_path.replace(final_path)
        return final_path


    def create_ping_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Alice Wonderland",
            from_email="alice@example.com",
            to_email="robot@company.local",
            subject="PING",
            body=(
                "Hello,\n\n"
                "I'm sending you a ping\n"
                "BR,\n" 
                "Alice"
            )
        )
        return self.write_eml_to_inbox(msg, prefix="ping")


    def create_job1_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Alice Wonderland",
            from_email="alice@example.com",
            to_email="robot@company.local",
            subject="Please run job1",
            body=(
                "I have no idea what job1 is though...\n"
                "Best regards,\n"
                "Alice\n"
            ),
            attachment_paths=[self.ATTACHMENTS_DIR / "job1_request.txt"],
        )
        return self.write_eml_to_inbox(msg, prefix="job1")

    def create_job1_b_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Bob Tester",
            from_email="bob@test.com",
            to_email="robot@company.local",
            subject="Job1",
            body=(
                "Hello,\n\n"
                "Please run job1\n\n"
                "order_number: 100245\n"
                "order_qty: 12000\n"
                "material_available: 11031\n\n"
                "Best regards,\n"
                "Bob\n"
            ),
            attachment_paths=[self.ATTACHMENTS_DIR / "job1_request.txt"],
        )
        return self.write_eml_to_inbox(msg, prefix="job1")


    def create_job2_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Bob Tester",
            from_email="bob@test.com",
            to_email="robot@company.local",
            subject="Job2 request",
            body=(
                "Hello,\n\n"
                "Please run job2 using attached file.\n\n"
                "Regards,\n"
                "Bob\n"
            ),
            attachment_paths=[self.ATTACHMENTS_DIR / "job2_request.csv"],
        )
        return self.write_eml_to_inbox(msg, prefix="job2")


    def create_unknown_job_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Charlie Strange",
            from_email="charlie@example.com",
            to_email="robot@company.local",
            subject="Do some weird magic",
            body=(
                "Hello,\n\n"
                "Please do that strange thing the robot probably cannot classify.\n\n"
                "Regards,\n"
                "Charlie\n"
            ),
        )
        return self.write_eml_to_inbox(msg, prefix="unknown")


    def create_blocked_sender_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Mallory Intruder",
            from_email="mallory@evil.com",
            to_email="robot@company.local",
            subject="Please run job1",
            body=(
                "Hello,\n\n"
                "I would like the robot to run job1.\n\n"
                "Regards,\n"
                "Mallory\n"
            ),
        )
        return self.write_eml_to_inbox(msg, prefix="blocked")


    def create_random_mail(self) -> Path:
        creators = [
            self.create_ping_mail,
            self.create_job1_mail,
            self.create_job1_b_mail,  # valid
            self.create_unknown_job_mail,
            self.create_blocked_sender_mail,
            self.create_job2_mail,
        ]
        return random.choice(creators)()


    def main(self) -> None:
        self.create_example_attachment_files()


class FakeQueryjobsGenerator:
    '''to create fake ERP jobs '''

    def add_random_row(self, path="Example_ERP_table.xlsx") -> str:

        if not os.path.isfile(path):
            raise RuntimeError("Example_ERP_table.xlsx not found, run main.py first ")
        wb = load_workbook(path)
        ws = wb.active

        assert ws is not None

        next_row = ws.max_row + 1

        erp_order_number = str(random.randint(10000000, 10999999))
        order_qty = random.randint(10, 100) * 100

        material_available = order_qty + random.randint(-100, 100)

        ws[f"A{next_row}"] = erp_order_number
        ws[f"B{next_row}"] = order_qty
        ws[f"C{next_row}"] = material_available

        wb.save(path)
        wb.close()
        return erp_order_number


class FakeJobsGenerator:
    ''' produce a fake email or a fake query-job at random'''

    def __init__(self) -> None:
        self.fake_emailjob = FakeEmailjobsGenerator()
        self.fake_queryjob = FakeQueryjobsGenerator()


    def run(self):
        while True:
            try:
                input("\nHit Enter to generate an random job")
                if random.randint(0,3) <= 2:
                    path = self.fake_emailjob.create_random_mail()
                    print(f"Created emailjob: {path.name}")
                else:
                    erp_order_number = self.fake_queryjob.add_random_row()
                    print(f"Created queryjob: {erp_order_number}")

            except KeyboardInterrupt:
                print("\nStopped.")
                break
            except Exception as err:
                print(f"WARN: generator error: {err}")
                time.sleep(1)


def main():

    if not os.path.isfile("main.py"):
        raise RuntimeError("Place this file in main.py directory")

    FakeJobsGenerator().run()


if __name__ == "__main__":
    main()