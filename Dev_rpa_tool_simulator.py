from __future__ import annotations
import random, os, threading, datetime, time, json, uuid, platform
from pathlib import Path
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from openpyxl import load_workbook #type: ignore

# to simualte the the behaviour of the RPA tool implementation 
class RPAToolSimulator:

    def run(self):

        self.log_system("RPAToolSimulator: i'm alive")
        print("RPAToolSimulator: i'm alive")

        while True:

            time.sleep(1)

            # read handover
            with open("handover.json", "r", encoding="utf-8") as f:
                handover_data = json.load(f)

            # claim workflow if "job_queued"
            ipc_state = handover_data.get("ipc_state")
            if ipc_state != "job_queued":
                continue

            # singal to Orchestrator the workflow is claimed 
            handover_data["ipc_state"] = "job_running"
            with open("handover.json", "w", encoding="utf-8") as f:
                json.dump(handover_data, f, indent=2)
            
            # identify job type
            job_type = handover_data.get("job_type")

            # ----------------------------------------------
            # JOB1
            # ----------------------------------------------
            if job_type == "job1":
                # retrive job-specific data  
                rpa_payload = handover_data.get("rpa_payload", {})
                erp_order_number = rpa_payload.get("source_ref")  
                new_qty = rpa_payload.get("target_order_qty")

                # simulation of job1 screenactiviy
                self.log_system(f"activities on screen_1 in ERP completed")
                self.log_system(f"activities on screen_2 in ERP completed")
                self.simulate_RPA_result_job1(erp_order_number, new_qty)

                # ready to verify result
                handover_data["ipc_state"] = "job_verifying"
            
            # ----------------------------------------------
            # JOB3
            # ----------------------------------------------
            elif job_type == "job3":
                # retrive job-specific data
                rpa_payload = handover_data.get("rpa_payload", {})
                erp_order_number = rpa_payload.get("source_ref")  
                new_qty = rpa_payload.get("target_order_qty")

                #simulation of job3 screenactiviy
                self.log_system(f"activities on screen_1 in ERP completed")
                self.log_system(f"activities on screen_2 in ERP completed")
                self.simulate_RPA_result_job1(erp_order_number, new_qty) # change to a "job3"-activity

                # ready to verify result
                handover_data["ipc_state"] = "job_verifying"

            # ----------------------------------------------
            # PING
            # ----------------------------------------------
            elif job_type == "ping":

                # play a sound
                if platform.system() == "Windows":
                    import winsound
                    winsound.Beep(1000, 300) #type: ignore

                elif platform.system() == "Linux":
                    print("\a", end="", flush=True)

                # ready to verify result
                handover_data["ipc_state"] = "job_verifying"
            
            # ----------------------------------------------
            # UNKOWN JOB
            # ----------------------------------------------
            else:
                self.log_system(f"no logic for job_type{job_type}")

                # error signal
                handover_data["ipc_state"] = "safestop"


            # ----------------------------------------------
            # Handover to Orchestrator
            # ----------------------------------------------
            with open("handover.json", "w", encoding="utf-8") as f:
                json.dump(handover_data, f, indent=2)

            self.log_system(f"RPASimulator.run() done, ipc_state: job_running -> job_verifying", handover_data.get("job_id"))


    def log_system(self, text: str, job_id=None):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
   
        job_part = f" | JOB {job_id}" if job_id else ""
        message = f"{timestamp} | RPA{job_part} | {text} \n"

        with open("system.log", "a", encoding="utf-8") as f:
            f.write(message)
            f.flush()
         


    

    def simulate_RPA_result_job1(self, erp_order_number: str, new_qty: int, path="Example_ERP_table.xlsx"):
        # here, updating a row.  IRL: updating in ERP

        assert erp_order_number is not None
        assert new_qty is not None

        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(erp_order_number):
                row[1].value = int(new_qty) #updating to the new value 'in ERP'     # type: ignore
                wb.save(path)
                wb.close()
                return True


def main():

    if not os.path.isfile("main.py"):
        raise RuntimeError("Place this file in main.py directory")

    rpa_tool_simulator = RPAToolSimulator()
    rpa_tool_simulator.run() #replace with RPA tool
    
main()