import atexit
import datetime
import json
import os
import platform
import re
import shutil
import signal
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import traceback
import tkinter as tk
from dataclasses import asdict, dataclass
from email import policy
from email.parser import BytesParser
from email.utils import parseaddr
from pathlib import Path
from typing import Any, Literal, Never, TypeAlias, get_args
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook  # type: ignore

VERSION = "0.4"

# =========================
# CUSTOMIZATION POINTS
# =========================
# Replace or adapt these parts for a real environment:
# - ExampleMailBackend
# - ExampleErpBackend
# - ExampleJob*Handler classes
# - NetworkService.NETWORK_HEALTHCHECK_PATH
# - FriendsRepository file format, if needed
# - the external RPA tool that reads/writes handover.json (not included here)


# ============================================================
# DATA MODELS
# ============================================================

HandoverState: TypeAlias = Literal["idle", "job_queued", "job_running", "job_verifying", "safestop"]
JobType: TypeAlias = Literal["ping", "job1", "job2", "job3", "job4"]
JobSourceType: TypeAlias = Literal["personal_inbox", "shared_inbox", "erp_query"]
JobStatus: TypeAlias = Literal["REJECTED", "QUEUED", "RUNNING", "VERIFYING", "FAIL", "DONE"]
JobAction: TypeAlias = Literal["DELETE_ONLY", "REPLY_AND_DELETE", "QUEUE_RPA_TOOL", "SKIP", "RETURN_TO_INBOX"]
UIStatus: TypeAlias = Literal["online", "safestop", "working", "no_network" , "out_of_office"]
ErrorCode: TypeAlias = Literal["PRE_HANDOVER_CRASH", "RPA_TOOL_CRASH", "VERIFICATION_MISMATCH", "POST_HANDOVER_CRASH", "OUT_OF_SERVICE","OUTSIDE_WORKING_HOURS", "UNKNOWN_JOB", "NO_ACCESS", "NO_NETWORK", "INVALID_INPUT", "CODE_ERROR", "RECOVERY_SOURCE_MISSING", "IN_SAFESTOP"]

@dataclass
class JobCandidate:
    source_ref: str
    job_source_type: JobSourceType
    source_data: dict[str, Any]

    email_address: str | None = None # for email only
    email_subject: str | None = None  # for email only
    email_body: str | None = None  # for email only

@dataclass
class JobDecision:
    action: JobAction
    job_type: JobType | None = None
    job_status: JobStatus | None = None
    error_code: ErrorCode | None = None
    error_message: str | None = None
    rpatool_payload: dict[str, Any] | None = None
    ui_log_message: str | None = None
    system_log_message: str | None = None
    send_online_notice: bool = False
    start_recording: bool = False

@dataclass
class HandoverJob:
    """Represents the payload stored in handover.json and exchanged with the RPA tool."""
    
    # common fields
    state: HandoverState

    source_ref: str | None = None  # identifier, eg. "ERP_ORDER:12345" or "mail1234.eml"

    job_type: JobType | None = None
    job_source_type: JobSourceType | None = None
    job_id: int | None = None

    email_address: str | None = None # for email
    email_subject: str | None = None      # for email
    email_body: str | None = None         # for email eg. "Hi, change the order 12345 to 44 pcs"

    # parsed from source 
    source_data: dict[str, Any] | None = None # eg. {"order_number": 12345, "target_qty": 44}

    # final instruction to RPA tool
    rpatool_payload: dict[str, Any] | None = None # eg. {"order_number": 12345, "target_qty": 44, "pick_qty_from_location": "WH7",} (saved in system.log, ensure no GDPR data)

@dataclass
class JobResult:
    is_success: bool
    error_message: str | None = None
    error_code: ErrorCode | None = None
    rpatool_payload: dict[str, Any] | None = None


class RuntimeFault(Exception):
    error_code = "UNSPEC_RUNTIME_FAULT"

    def __init__(self, message:str, job_id:int|None=None, handover_job:HandoverJob|None=None, cause:Exception|None=None, traceback_text:str|None=None):
        super().__init__(message)
        self.error_message = message
        self.job_id = job_id
        self.handover_job = handover_job
        self.cause = cause
        self.traceback_text = traceback_text
        self.error_code = self.__class__.error_code

class PreHandoverCrash(RuntimeFault):
    error_code = "PRE_HANDOVER_CRASH"
class RpaToolCrash(RuntimeFault):
    error_code = "RPA_TOOL_CRASH"
class VerificationMismatch(RuntimeFault):
    error_code = "VERIFICATION_MISMATCH"
class PostHandoverCrash(RuntimeFault):
    error_code = "POST_HANDOVER_CRASH"

# ============================================================
# EXAMPLE BACKENDS
# ============================================================

class ExampleMailBackend:
    """
    Example mailbox backend that simulates mailbox processing using local
    folders and .eml files.

    Replace this with a real backend, for example Outlook or Microsoft Graph.
    """

    def __init__(self, logger, job_source_type) -> None:
        self.logger = logger
        self.job_source_type = job_source_type # change to folder in e.g. outlook
        self.inbox_dir = Path(self.job_source_type) / "inbox"
        self.processing_dir = Path(self.job_source_type) / "processing"

        self.inbox_dir.mkdir(parents=True, exist_ok=True)
        self.processing_dir.mkdir(parents=True, exist_ok=True)


    def list_inbox_mail_paths(self, max_items=None) -> list[str]:
        paths_raw = sorted(self.inbox_dir.glob("*.eml"))

        if max_items is not None:
            paths_raw = paths_raw[:max_items]

        paths = [str(x) for x in paths_raw] #convert Path-type to str
        return paths


    def list_processing_mail_paths(self, max_items=None) -> list[str]:
        paths_raw = sorted(self.processing_dir.glob("*.eml"))

        if max_items is not None:
            paths_raw = paths_raw[:max_items]

        paths = [str(x) for x in paths_raw] #convert Path-type to str
        return paths
    

    def parse_mail_file(self, mail_path) -> JobCandidate:
        with open(mail_path, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)

        from_name, from_address = parseaddr(msg.get("From", ""))
        del from_name # not used

        email_address = (from_address or "").strip().lower()
        if not email_address or "@" not in email_address:
            email_address = None

        email_subject = msg.get("Subject", "").strip()

        

        # message_id = msg.get("Message-ID", "").strip()
        # not needed. source_ref is sufficient (in this example: Path.   In outlook: Outlook EntryID / Graph ID)

        # raw_headers = {k: str(v) for k, v in msg.items()}   
        # not needed (but good for troubleshooting all metadata) 

        if msg.is_multipart():
            body_parts = []
            for part in msg.walk():
                if part.get_content_type() == "text/plain" and not part.get_filename():
                    try:
                        body_parts.append(part.get_content())
                    except Exception:
                        pass
            email_body = "\n".join(body_parts).strip()
        else:
            try:
                email_body = msg.get_content().strip()
            except Exception:
                email_body = ""
        

        # placeholder for implementation
        attachments = {}
        #attachments = {
        #    "attachments": [
        #        {
        #            "filename": "orders.xlsx",
        #            "path": "/some/path/orders.xlsx",
        #        }
        #    ]
        #}
     
        return JobCandidate(
            source_ref=mail_path,
            email_address=email_address,
            email_subject=email_subject,
            email_body=email_body,
            job_source_type=self.job_source_type,
            source_data=attachments,
            )


    def move_to_processing(self, mail: JobCandidate) -> JobCandidate:

        example_path = Path(mail.source_ref) # example backend use Path

        target_path = self.processing_dir / example_path.name # .name for filename only
        shutil.move(str(example_path), str(target_path))
        
        self.logger.system(f"moved {example_path} to {target_path}")
        mail.source_ref = str(target_path)

        return mail
        

    def reply_and_delete(self, candidate: JobCandidate, extra_subject: str, extra_body: str, job_id: int) -> None:
        self.send_reply(candidate, extra_subject, extra_body, job_id)
        self.delete_from_processing(candidate, job_id)


    def send_reply(self, candidate: JobCandidate, extra_subject: str, extra_body: str, job_id: int) -> None:

        reply_to = candidate.email_address
        subject = f"{extra_subject} re: {candidate.email_subject}"
        body = (
            f"{extra_body} \n\n"
            f"-------------------------------------------------------------\n"
            f"{candidate.email_body}"
        ) # In a real mail backend, this should use the provider's native reply mechanism.

        reply_message = f"reply_to={reply_to}, subject={subject}, body={body}"
        self.logger.system(reply_message[:200], job_id)
        
        assert reply_to is not None
        self._print_email_preview(reply_to, subject, body)


    def _print_email_preview(self, reply_to: str, subject: str, body: str):

        print(
        "\n" + "="*72 +
        "\n📧 EMAIL REPLY PREVIEW\n" +
        "="*72 +
        f"\nFrom:    robot@runtime.local"
        f"\nTo:      {reply_to}"
        f"\nSubject: {subject}"
        f"\nDate:    {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        "\n" + "-"*72 +
        f"\n{body}\n" +
        "="*72 + "\n"
    )
        

    def delete_from_processing(self, candidate: JobCandidate, job_id: int | None = None) -> None:

        self.logger.system(f"removing: {candidate.source_ref}", job_id)
        os.remove(candidate.source_ref)


    def return_to_inbox(self, candidate: JobCandidate, job_id: int) -> None:
        ''' to simplify for end-user, return unhandled emails to origin location'''
        
        # placeholder for implementation

        # rename email subject to "FAIL/" and ignore renamed obj. in _is_shared_inbox_email_in_scope()

        example_path = Path(candidate.source_ref)

        target_path = self.inbox_dir / example_path.name #.name only the filename
        shutil.move(str(example_path), str(target_path))
        
        self.logger.system(f"moved {candidate} back to inbox", job_id)


class ExampleErpBackend:
    """Example ERP backend backed by a local Excel file."""
    def select_mismatch_rows(self, path="Example_ERP_table.xlsx") -> list[dict]:
        # do a well targeted 'query' 

        self._ensure_example_erp_exists(path)

        try:
            wb = load_workbook(path)
        except BadZipFile:
            time.sleep(1)
            wb = load_workbook(path)

        ws = wb.active

        assert ws is not None #to satisfy pylance

        all_rows=[]

        for row in ws.iter_rows(min_row=2):  # skip header
            
            source_ref = row[0].value
            order_qty = row[1].value
            material_available = row[2].value

            if order_qty != material_available:

                all_rows.append({
                        "source_ref": source_ref,
                        "order_qty": order_qty,
                        "material_available": material_available,
                    })
                
        wb.close()
        return all_rows
    
    
    def build_candidate_from_row(self, row) -> JobCandidate:
              
        source_ref = row.get("source_ref")
        order_qty = row.get("order_qty")
        material_available = row.get("material_available")


        try: order_qty = int(order_qty)
        except Exception: raise ValueError(f"invalid order_qty: {order_qty}")
        try: material_available = int(material_available)
        except Exception: raise ValueError(f"invalid material_available: {material_available}")


        source_data ={
            "order_qty": order_qty,
            "material_available": material_available,
        }

        return JobCandidate(
            source_ref=str(source_ref),
            job_source_type="erp_query",
            source_data=source_data
        )

    
    def _ensure_example_erp_exists(self, path="Example_ERP_table.xlsx") -> None:
        """Create the demo ERP table if it does not exist."""
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None #to satisfy pylance

        # headers
        ws["A1"] = "source_ref"
        ws["B1"] = "order_qty"
        ws["C1"] = "material_available"

        wb.save(path)
        wb.close()

    
    def get_order_qty(self, source_ref, path="Example_ERP_table.xlsx") -> int | None:
        self._ensure_example_erp_exists(path)

        try:
            wb = load_workbook(path)
        except BadZipFile:
            time.sleep(1)
            wb = load_workbook(path)

        ws = wb.active
        assert ws is not None #to satisfy pylance

        for row in ws.iter_rows(min_row=2):
            cell_source_ref = row[0].value

            if str(cell_source_ref) == str(source_ref):
                value = row[1].value  # order_qty

                if isinstance(value, int):
                    wb.close()
                    return int(value)
                
                else: 
                    raise ValueError(f"order_qty: {value} is not INT")
        
        wb.close()
        return None


# ============================================================
# JOB FLOWS
# ============================================================

class MailFlow:
    """Handle email-driven job intake."""
    def __init__(self, logger, friends_repo, is_within_operating_hours, network_service, job_handlers, pre_handover_service, mail_backend_personal, mail_backend_shared) -> None:
        self.logger = logger
        self.friends_repo = friends_repo
        self._is_within_operating_hours = is_within_operating_hours
        self.network_service = network_service
        self.job_handlers = job_handlers
        self.pre_handover_service = pre_handover_service
        self.mail_backend_personal = mail_backend_personal
        self.mail_backend_shared = mail_backend_shared


    def poll_once(self) -> bool:
        ''' a candidate is any email from personal_inbox OR an 'in scope'-email from shared_inbox '''

        # claim and parse from all mail-sources
        candidate = self._claim_next_mail_candidate() 
        if not candidate:
            return False
        
        
        # personal inbox = direct human-to-runtime channel 
        if candidate.job_source_type == "personal_inbox":
            self.friends_repo.reload_if_modified()
            self.logger.ui(f"email from {candidate.email_address}", blank_line_before=True)
            decision = self._decide_personal_inbox_email(candidate)       

        # shared inbox = external business mailbox
        elif candidate.job_source_type == "shared_inbox":
            decision = self._decide_shared_inbox_email(candidate)

        else:
            raise ValueError(f"unknown job_source_type={candidate.job_source_type}")

        self.pre_handover_service.apply_decision(candidate, decision)
        return True


    def _claim_next_mail_candidate(self) -> JobCandidate | None:

        # personal inbox (parse, always claim)
        paths = self.mail_backend_personal.list_inbox_mail_paths(max_items=1)
       
        for path in paths:
            mail = self.mail_backend_personal.parse_mail_file(path)
            mail = self.mail_backend_personal.move_to_processing(mail)
            self.logger.system(f"{mail.job_source_type} produced mail {mail.source_ref}")
            return mail
        
        # shared inbox (parse, maybe claim)
        paths = self.mail_backend_shared.list_inbox_mail_paths()
        
        for path in paths:
            mail = self.mail_backend_shared.parse_mail_file(path)

            if not self._is_shared_inbox_email_in_scope(mail):
                continue
            
            mail = self.mail_backend_shared.move_to_processing(mail)
            self.logger.system(f"{mail.job_source_type} produced mail {mail.source_ref}")

            return mail

        return None


    def _is_shared_inbox_email_in_scope(self, mail: JobCandidate) -> bool:
  
        # Intentionally minimal example.
        self.logger.system(f"checking sender: {mail.email_address} email_subject: {mail.email_subject}")

        # skip emails moved back by return_to_inbox()
        if str(mail.email_subject).upper().startswith("FAIL/"):
            return False
        
        # Placeholder for mailbox-specific scope rules, eg. email-adress or subject matching.
        
        return True

    
    def _classify_personal_inbox_email(self, mail: JobCandidate) -> JobType | None:

        email_subject = str(mail.email_subject).strip().lower()

        if email_subject == "ping":
            return "ping"
        
        elif "job1" in email_subject.lower():
            return"job1"
        
        elif "job2" in email_subject.lower():
            return "job2"

        return None


    def _classify_shared_inbox_email(self):
        # placeholder for implementation
        pass
 

    def _decide_personal_inbox_email(self, mail: JobCandidate) -> JobDecision:
        '''decide what to do with the found candidate'''
        precheck_result: JobResult
        job_type = None

        try:
            if not self.friends_repo.is_allowed_sender(mail.email_address):
                return JobDecision(
                    action="DELETE_ONLY",
                    ui_log_message="--> rejected (not in friends.xlsx)",
                    system_log_message="--> rejected (not in friends.xlsx)"
                )

            if not self._is_within_operating_hours():
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_status="REJECTED",
                    error_code="OUTSIDE_WORKING_HOURS",
                    error_message="Outside robot's working hours 05-23.",
                    ui_log_message="--> rejected (outside working hours)",
                    system_log_message="--> rejected (outside working hours)",
                )

            job_type = self._classify_personal_inbox_email(mail)

            if job_type is None:
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=None,
                    job_status="REJECTED",
                    error_code="UNKNOWN_JOB",
                    error_message="Could not identify a job type from your email, check spelling and/or name for attached file(s).",
                    ui_log_message="--> rejected (unable to identify job type)",
                    system_log_message="--> rejected (unable to identify job type)",
                )
            
            if not self.friends_repo.has_job_access(mail.email_address, job_type):
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="NO_ACCESS",
                    error_message=f"Request denied, your email is not permitted to trigger '{job_type}'. Contact the robot administrator.",
                    ui_log_message=f"--> rejected (no access to {job_type})",
                    system_log_message=f"--> rejected (no access to {job_type})",
                )

            if not self.network_service.has_network_access():
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="NO_NETWORK",
                    error_message="Robot has no network connection at the moment.",
                    ui_log_message="--> rejected (no network connection)",
                    system_log_message="--> rejected (no network connection)",
                )

            handler = self.job_handlers.get(job_type)
            if handler is None:
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_status="REJECTED",
                    job_type=job_type,
                    error_code="PRE_HANDOVER_CRASH",
                    error_message=f"No handler registered for job_type={job_type}",
                )

            precheck_result = handler.precheck_and_build_payload(mail)
            if not precheck_result.is_success:
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="INVALID_INPUT",
                    error_message=precheck_result.error_message,
                    ui_log_message=f"--> rejected (invalid input for {job_type})",
                    system_log_message=f"--> rejected (invalid input for {job_type})",
                )
            
            return JobDecision(
                action="QUEUE_RPA_TOOL",
                job_type=job_type,
                job_status="QUEUED",
                system_log_message=f"accepted ({job_type})",
                send_online_notice=True,
                start_recording=True,
                rpatool_payload=precheck_result.rpatool_payload
            )

        except Exception as err:
            # this branch will reject due to PRE_HANDOVER_CRASH (improve to also notify admin? change to crash-path)
            return JobDecision(
                action="REPLY_AND_DELETE",
                job_status="REJECTED",
                job_type=None,
                ui_log_message=f"--> rejected (code error in job_type {job_type})",
                system_log_message=f"--> rejected (code error in job_type {job_type})",
                error_code="PRE_HANDOVER_CRASH",
                error_message=f"unhandled error: {err}",
            )
    

    def _decide_shared_inbox_email(self, mail: JobCandidate) -> JobDecision:
        ''' decide what to do with the found candidate '''
        # placeholder for implementation

        return JobDecision(
            action="RETURN_TO_INBOX",
            job_type=None,
            error_code="CODE_ERROR",
            error_message="No logic implemented yet to handle shared_inbox mails",
            system_log_message="No logic implemented yet to handle shared_inbox mails",
        )


class QueryFlow:
    """Handle query-driven job intake."""
    def __init__(self, logger, audit_repo, job_handlers, pre_handover_service, is_within_operating_hours, erp_backend) -> None:
        self.logger = logger
        self.audit_repo = audit_repo
        self.job_handlers = job_handlers
        self.pre_handover_service = pre_handover_service
        self._is_within_operating_hours = is_within_operating_hours
        self.erp_backend = erp_backend


    def poll_once(self) -> bool:
        ''' find candidates in form of a query row '''
        
        if not self._is_within_operating_hours():
            return False

        # check if new match
        candidate = self._fetch_next_query_candidate()
        if not candidate:
            return False

        self.logger.ui(f"query job detected for {candidate.source_ref}", blank_line_before=True)
        
        # decide what to do
        decision = self._decide_candidate(candidate)
        
        # do it
        self.pre_handover_service.apply_decision(candidate, decision)
        return True
    
        
    def _fetch_next_query_candidate(self) -> JobCandidate | None:

        # job 3 example
        all_selected_rows = self.erp_backend.select_mismatch_rows()
        
        if not all_selected_rows:
            return None
    
        for row_candidate_raw in all_selected_rows:
            row_candidate = self.erp_backend.build_candidate_from_row(row_candidate_raw)

            # Avoid reprocessing the same source_ref multiple times on the same day.
            if self.audit_repo.has_been_processed_today(row_candidate.source_ref):
                continue

            row_candidate.job_source_type="erp_query"
            self.logger.system(f"{row_candidate.job_source_type} produced source_ref {row_candidate.source_ref}")
            return row_candidate
        
        # job 4
        # placeholder for implementation
        
        return None


    def _decide_candidate(self, candidate: JobCandidate) -> JobDecision:
        precheck_result: JobResult
        self.logger.system("running")

        job_type = None

        try:

            # placeholder for implementation
            # eg. below:


            job_type = self._classify_candidate(candidate)
            handler = self.job_handlers.get(job_type)

            if handler is None:
                return JobDecision(
                    action="SKIP",
                    job_status="REJECTED",
                    job_type=job_type,
                    error_code="PRE_HANDOVER_CRASH",
                    error_message=f"No handler found for job_type={job_type}",
                )
            
            # do the precheck from "job specifics"-section
            precheck_result = handler.precheck_and_build_payload(candidate)

            if not precheck_result.is_success:
                return JobDecision(
                    action="SKIP",  
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="INVALID_INPUT",
                    error_message=precheck_result.error_message,
                    ui_log_message=f"--> rejected (invalid input for {job_type})",
                )
        
            return JobDecision(
                action="QUEUE_RPA_TOOL",
                job_type=job_type,
                job_status="QUEUED",
                system_log_message=f"accepted ({job_type})",
                start_recording=True,
                rpatool_payload=precheck_result.rpatool_payload,
                )

        except Exception as err:
            return JobDecision(
                action="SKIP",
                job_status="REJECTED",
                job_type=job_type,
                error_code="PRE_HANDOVER_CRASH",
                error_message=str(err),
            )
    

    def _classify_candidate(self, candidate: JobCandidate) -> JobType | None:
        # placeholder for implementation

        return "job3"


class PreHandoverService:
    """Execute pre-handover actions and build HandoverJob objects for the RPA tool."""
    def __init__(self, logger, handover_repo, update_ui_status, show_recording_overlay, generate_job_id, recording_service, audit_repo, notification_service, mail_backend_personal, mail_backend_shared) -> None:
        self.logger = logger
        self.handover_repo = handover_repo
        self.recording_service = recording_service
        self._generate_job_id = generate_job_id
        self.audit_repo = audit_repo
        self._update_ui_status = update_ui_status
        self._show_recording_overlay = show_recording_overlay
        self.notification_service = notification_service
        self.mail_backend_personal = mail_backend_personal
        self.mail_backend_shared = mail_backend_shared
    

    def _validate_decision(self, decision: JobDecision) -> None:

        if not isinstance(decision, JobDecision):
            raise ValueError("decision must be JobDecision")

        # validate action
        if decision.action not in get_args(JobAction):
            raise ValueError(f"invalid action: {decision.action}")

        # validate job_type
        if decision.job_type is not None and decision.job_type not in get_args(JobType):
            raise ValueError(f"invalid job_type: {decision.job_type}")

        # validate job_status
        if decision.job_status is not None and decision.job_status not in get_args(JobStatus):
            raise ValueError(f"invalid job_status: {decision.job_status}")

        # ============================================================
        # ACTION-SPECIFIC RULES
        # ============================================================

        action = decision.action

        # ------------------------------------------------------------
        # DELETE_ONLY
        # ------------------------------------------------------------
        if action == "DELETE_ONLY":
            if decision.job_status is not None:
                raise ValueError("DELETE_ONLY must not have job_status")
            if decision.rpatool_payload is not None:
                raise ValueError("DELETE_ONLY must not have rpatool_payload")

        # ------------------------------------------------------------
        # REPLY_AND_DELETE
        # ------------------------------------------------------------
        elif action == "REPLY_AND_DELETE":
            if decision.job_status != "REJECTED":
                raise ValueError("REPLY_AND_DELETE requires job_status REJECTED")

            if decision.error_message is None:
                raise ValueError("REPLY_AND_DELETE requires error_message (reply text)")

            if decision.rpatool_payload is not None:
                raise ValueError("REPLY_AND_DELETE must not have rpatool_payload")
   
        # ------------------------------------------------------------
        # QUEUE_RPA_TOOL
        # ------------------------------------------------------------
        elif action == "QUEUE_RPA_TOOL":
            if decision.job_type is None:
                raise ValueError("QUEUE_RPA_TOOL requires job_type")

            if decision.job_status != "QUEUED":
                raise ValueError("QUEUE_RPA_TOOL requires job_status='QUEUED'")

            if decision.rpatool_payload is None:
                raise ValueError("QUEUE_RPA_TOOL requires rpatool_payload")
            

            if not isinstance(decision.rpatool_payload, dict):
                raise ValueError("rpatool_payload must be dict")

        # ------------------------------------------------------------
        # SKIP (query flow reject)
        # ------------------------------------------------------------
        elif action == "SKIP":
            if decision.job_status != "REJECTED":
                raise ValueError("SKIP requires job_status='REJECTED'")

        # ------------------------------------------------------------
        # RETURN_TO_INBOX
        # ------------------------------------------------------------
        elif action == "RETURN_TO_INBOX":
            if decision.job_status is not None:
                raise ValueError("RETURN_TO_INBOX should not set job_status")

        else:
            # should never happen due to earlier validation
            raise ValueError(f"Unhandled action: {action}")
        

    def _validate_candidate_decision_combination(self, candidate: JobCandidate, decision: JobDecision) -> None:

        if candidate.job_source_type not in get_args(JobSourceType):
            raise ValueError(f"invalid candidate.job_source_type: {candidate.job_source_type}")

        is_mail = candidate.job_source_type in ("personal_inbox", "shared_inbox")
        is_query = candidate.job_source_type == "erp_query"

        # ------------------------------------------------------------
        # source-type-specific candidate sanity
        # ------------------------------------------------------------
        if is_mail:
            if candidate.email_address is None:
                raise ValueError(f"{candidate.job_source_type} candidate requires email_address")
            if candidate.email_subject is None:
                raise ValueError(f"{candidate.job_source_type} candidate requires email_subject")
            if candidate.email_body is None:
                raise ValueError(f"{candidate.job_source_type} candidate requires body")

        elif is_query:
            if candidate.source_data is None:
                raise ValueError("erp_query candidate requires source_data")

        else:
            raise ValueError(f"unknown candidate.job_source_type: {candidate.job_source_type}")

        # ------------------------------------------------------------
        # action must match candidate source type
        # ------------------------------------------------------------
        if decision.action in ("DELETE_ONLY", "REPLY_AND_DELETE", "RETURN_TO_INBOX"):
            if not is_mail:
                raise ValueError(
                    f"action {decision.action} is only valid for mail candidates, not {candidate.job_source_type}"
                )

        if decision.action == "SKIP":
            if not is_query:
                raise ValueError(
                    f"action SKIP is only valid for query candidates, not {candidate.job_source_type}"
                )

        # ------------------------------------------------------------
        # optional policy checks
        # ------------------------------------------------------------
        if decision.send_online_notice:
            if candidate.job_source_type != "personal_inbox":
                raise ValueError("lifesign only valid for personal_inbox")

            if not candidate.email_address:
                raise ValueError("lifesign requires email_address")
            
        # RETURN_TO_INBOX is for shared-only emails
        if decision.action == "RETURN_TO_INBOX":
            if candidate.job_source_type != "shared_inbox":
                raise ValueError("RETURN_TO_INBOX is only valid for shared_inbox")

        # DELETE_ONLY is for personal-only emails.
        if decision.action == "DELETE_ONLY":
            if candidate.job_source_type != "personal_inbox":
                raise ValueError("DELETE_ONLY is only valid for personal_inbox")


        # REPLY_AND_DELETE should be used for personal inbox only 
        if decision.action == "REPLY_AND_DELETE" and candidate.job_source_type != "personal_inbox":
            raise ValueError("REPLY_AND_DELETE is only valid for personal_inbox")

        # send_online_notice only makes sense together with QUEUE_RPA_TOOL
        if decision.send_online_notice and decision.action != "QUEUE_RPA_TOOL":
            raise ValueError("send_online_notice requires action='QUEUE_RPA_TOOL'")

        # start_recording only makes sense for queued jobs
        if decision.start_recording and decision.action != "QUEUE_RPA_TOOL":
            raise ValueError("start_recording requires action='QUEUE_RPA_TOOL'")


    def _log_decision_messages(self, decision: JobDecision):
        if decision.ui_log_message:
            self.logger.ui(decision.ui_log_message)

        if decision.system_log_message:
            self.logger.system(decision.system_log_message)


    def _maybe_send_online_notice(self, candidate: JobCandidate, decision: JobDecision, job_id: int|None):
        if decision.send_online_notice and not self.audit_repo.has_sender_job_today(candidate.email_address, job_id):
            self.notification_service.send_online_notice(candidate, job_id)


    def _maybe_start_recording(self, decision: JobDecision, job_id: int|None):
        
        if decision.start_recording:
            self.recording_service.start(job_id)
            
            try: self._show_recording_overlay()
            except Exception as e: self.logger.system(f"error {e}", job_id)


    def _insert_audit_row(self, candidate: JobCandidate, decision: JobDecision, job_id: int|None,):
        now = datetime.datetime.now()
        job_finish_time = None if decision.action == "QUEUE_RPA_TOOL" else now.strftime("%H:%M:%S")

        self.audit_repo.insert_job(
        job_id=job_id,
        source_ref=candidate.source_ref,
        email_address=candidate.email_address,
        email_subject=candidate.email_subject,
        job_type=decision.job_type,
        job_start_date=now.strftime("%Y-%m-%d"),
        job_start_time=now.strftime("%H:%M:%S"),
        job_finish_time=job_finish_time,
        job_status=decision.job_status,
        job_source_type=candidate.job_source_type,
        error_code=decision.error_code,
        error_message=decision.error_message,
        )


    def _build_handover_job(self, candidate: JobCandidate, decision: JobDecision, job_id: int| None) -> HandoverJob:
        
        return HandoverJob(
            state="job_queued",
            job_id=job_id,
            job_type=decision.job_type,
            job_source_type=candidate.job_source_type,
            source_ref=candidate.source_ref,
            email_address=candidate.email_address,
            email_subject=candidate.email_subject,
            email_body=candidate.email_body,
            source_data=candidate.source_data,
            rpatool_payload=decision.rpatool_payload,
            )
    

    def apply_decision(self, candidate: JobCandidate, decision: JobDecision) -> None:
        job_id = None
        
        try:
            self._validate_decision(decision)
            self._validate_candidate_decision_combination(candidate, decision)

            self._log_decision_messages(decision)
            
            # DELETE_ONLY is intentionally non-audited.
            if decision.action == "DELETE_ONLY":
                self.mail_backend_personal.delete_from_processing(candidate)
                return
            

            job_id = self._generate_job_id()
            
            # prioritize new audit row
            self._insert_audit_row(candidate, decision, job_id,)

            # REPLY_AND_DELETE
            if decision.action == "REPLY_AND_DELETE": # only personal-inbox
                self.notification_service.send_final_reply_and_delete_original(
                    candidate=candidate,
                    job_status="REJECTED",
                    job_id=job_id,
                    error_code=decision.error_code,
                    reason=decision.error_message,
                )
                self.audit_repo.update_job(job_id=job_id, final_reply_sent=True,)
                return
            
            # SKIP
            if decision.action == "SKIP":
                return

            # RETURN_TO_INBOX
            elif decision.action == "RETURN_TO_INBOX": # only shared-inbox e.g. error with email in scope
                self.mail_backend_shared.return_to_inbox(candidate, job_id)
                return
                            
            # QUEUE_RPA_TOOL
            if decision.action == "QUEUE_RPA_TOOL":
                self._update_ui_status(forced_status="working")
                handover_job = self._build_handover_job(candidate, decision, job_id)
                self._maybe_start_recording(decision, job_id)
                self._maybe_send_online_notice(candidate, decision, job_id)
                self.handover_repo.write(handover_job)
                return

        except Exception as e:
            raise PreHandoverCrash(str(e), job_id=job_id, cause=e) from e


class PostHandoverService:
    """Finalize and verify a handover job returned by the RPA tool."""
    def __init__(self, logger, audit_repo, job_handlers, recording_service, hide_recording_overlay, mail_backend_personal, mail_backend_shared, notification_service) -> None:
        self.logger = logger
        self.audit_repo = audit_repo
        self.job_handlers = job_handlers
        self.recording_service = recording_service
        self._hide_recording_overlay = hide_recording_overlay
        self.mail_backend_personal = mail_backend_personal
        self.mail_backend_shared = mail_backend_shared
        self.notification_service = notification_service


    def finalize_handover_job(self, handover_job: HandoverJob) -> None:
        '''
        verify_result() must return:
        * success, or
        * failure with VERIFICATION_MISMATCH
        all other outcomes are treated as programming/system fault. (implement eg. QUERY_TIMEOUT if needed)
        '''
        job_id = handover_job.job_id
        job_type = handover_job.job_type

        try:
            # Mark that runtime has reclaimed ownership of workflow
            self.audit_repo.update_job(job_id=job_id, job_status="VERIFYING")
            
            handler = self.job_handlers.get(job_type)
            if handler is None:
                raise PostHandoverCrash(
                    f"No handler for job_type={job_type}",
                    job_id=job_id,
                    handover_job=handover_job,
                )
            
            verification_result = handler.verify_result(handover_job)

            if verification_result.is_success:
                self._finalize_job_result(handover_job, job_status="DONE")
                return

            # consider mismatch as a critical error 
            if verification_result.error_code == "VERIFICATION_MISMATCH":
                error_message = str(verification_result.error_message)
                self._finalize_job_result(
                    handover_job,
                    job_status="FAIL",
                    error_code=verification_result.error_code,
                    error_message=error_message,
                )
                raise VerificationMismatch(
                    error_message,
                    job_id=job_id,
                    handover_job=handover_job,
                )
            
            raise PostHandoverCrash(
                f"Unknown verify_result outcome for job_type={job_type}: "
                f"error_code={verification_result.error_code}, "
                f"error_message={verification_result.error_message}",
                job_id=job_id,
                handover_job=handover_job,
            )
            
        except RuntimeFault:
            raise
            
        except Exception as err:
            try:
                self.audit_repo.update_job(
                    job_id=job_id,
                    job_status="FAIL",
                    error_code="POST_HANDOVER_CRASH",
                    error_message=f"crash during verification stage: {err}",
                    job_finish_time=datetime.datetime.now().strftime("%H:%M:%S"),
            )
            except Exception as err2:
                self.logger.system(f"[PostHandoverService] {err} {err2}", job_id)
            
            raise PostHandoverCrash(
                f"verification stage crashed, outcome unknown: {err}",
                job_id=job_id,
                handover_job=handover_job,
                cause=err,
            ) from err


    def _update_audit(self, job_id, job_status, error_code, jobhandler_error_message,) -> None:
        
        now = datetime.datetime.now().strftime("%H:%M:%S")
        
        self.audit_repo.update_job(
            job_id=job_id, 
            job_status=job_status, 
            error_code=error_code, 
            error_message=jobhandler_error_message, 
            job_finish_time=now,
            )


    def _update_logs(self, job_status: str, handover_job: HandoverJob,) -> None:
        
        job_status = job_status.lower()
        job_type = handover_job.job_type

        # ui/dashboard log
        self.logger.ui(f"--> {job_status} ({job_type})")

        # system log (system.log)
        self.logger.system(f"{job_status} ({job_type})", handover_job.job_id)

        
    def _finalize_job_result(self, handover_job: HandoverJob, job_status: JobStatus, error_code: str | None=None, error_message: str | None=None):
        job_id = handover_job.job_id

        # update audit w/ result (DONE/FAIL)
        self._update_audit(job_id, job_status, error_code, error_message)
        
        # do side effects
        self.recording_service.stop(job_id)
        self.recording_service.upload_recording(job_id)
        self._hide_recording_overlay()
 
        # do mail source specifics
        final_reply_sent = self._handle_source_completion(handover_job, job_status, error_code, error_message)

        if final_reply_sent:
            self.audit_repo.update_job(job_id=job_id, final_reply_sent=True,)

        # update ui w/ result (DONE/FAIL)
        self._update_logs(job_status, handover_job)

     
    def _map_candidate_from_handover_job(self, handover_job: HandoverJob) -> JobCandidate:
        assert handover_job.source_ref is not None # to satisfy pylance
        assert handover_job.source_data is not None # to satisfy pylance
        assert handover_job.job_source_type is not None # to satisfy pylance
        
        return JobCandidate(
                source_ref=handover_job.source_ref,
                job_source_type=handover_job.job_source_type,
                source_data=handover_job.source_data,
                email_address=handover_job.email_address,
                email_subject=handover_job.email_subject,
                email_body=handover_job.email_body,
                )
        
  
    def _handle_source_completion(self, handover_job: HandoverJob, job_status: JobStatus, error_code: str | None, error_message: str | None) -> bool:

        if handover_job.job_source_type not in ("personal_inbox", "shared_inbox"):
            return False
    
        # rebuild candidate (from rebuilt handover_job)
        candidate = self._map_candidate_from_handover_job(handover_job)

        job_id = handover_job.job_id

        if handover_job.job_source_type == "personal_inbox":

            if job_status == "DONE":
                self.notification_service.send_final_reply_and_delete_original(
                    candidate=candidate,
                    job_status=job_status,
                    job_id=job_id,
                    error_code=None,
                )
                return True

            elif job_status == "FAIL":
                self.notification_service.send_final_reply_and_delete_original(
                    candidate=candidate,
                    job_status=job_status,
                    error_code=error_code,
                    job_id=job_id,
                    reason=error_message,
                )
                return True

            raise ValueError(f"unexpected job_status in _handle_source_completion(): {job_status}")

        # delete shared mail (or move to archive?)
        if handover_job.job_source_type == "shared_inbox":
            self.mail_backend_shared.delete_from_processing(candidate, job_id)
            return False
        
        return False


# ============================================================
# JOB SPECIFICS
# ============================================================

class ExampleJob1Handler:
    ''' everything for job1 '''
    def __init__(self, logger) -> None:
        self.logger = logger


    def precheck_and_build_payload(self, candidate: JobCandidate) -> JobResult:
        ''' sanity-check (and ERP check) on given data '''
        email_body = candidate.email_body
        assert email_body is not None # to satisfy pylance

        # get important info for job1, eg.:
        order_number_match = re.search(r"order_number:\s*(.+)", email_body)
        order_number = order_number_match.group(1) if order_number_match else None

        order_qty_match = re.search(r"order_qty:\s*(.+)", email_body)
        order_qty = order_qty_match.group(1) if order_qty_match else None

        material_available_match = re.search(r"material_available:\s*(.+)", email_body)
        material_available = material_available_match.group(1) if material_available_match else None

        error_message = ""
        if order_number is None:
            error_message += "missing order_number. "
        if order_qty is None:
            error_message += "missing order_qty. "
        if material_available is None:
            error_message += "missing material_available. "

        if error_message:
            return JobResult(is_success=False, error_message=error_message.strip())

        # and for any attachments, eg:
        attachments = candidate.source_data.get("attachments", [])
        #for attachment in attachments:
        #    print(attachment.get("filename"))

        rpatool_payload = {
            "order_number": order_number,
            "order_qty": order_qty,
            "target_order_qty": material_available,
            "attachments": attachments,
        }

        return JobResult(is_success=True, rpatool_payload=rpatool_payload)
    

    def verify_result(self, handover_job: HandoverJob) -> JobResult:
        return JobResult(is_success=True)


class ExampleJob2Handler:
    ''' everything for job2 '''
    def __init__(self, logger) -> None:
        self.logger = logger
   

    def precheck_and_build_payload(self, candidate: JobCandidate) -> JobResult:
        # placeholder for implementation

        return JobResult(is_success=False, error_message="no logic for job2.")


    def verify_result(self, handover_job: HandoverJob) -> JobResult:
        return JobResult(is_success=True)


class ExamplePingJobHandler:
    ''' everything for ping '''
    def __init__(self, logger) -> None:
        self.logger = logger


    def precheck_and_build_payload(self, candidate: JobCandidate) -> JobResult:
        return JobResult(is_success=True, rpatool_payload={})


    def verify_result(self, handover_job: HandoverJob) -> JobResult:
        return JobResult(is_success=True)
    
   
class ExampleJob3Handler:
    ''' everything for job3 '''
    def __init__(self, logger, erp_backend) -> None:
        self.logger = logger
        self.erp_backend = erp_backend

   
    def precheck_and_build_payload(self, candidate: JobCandidate) -> JobResult:
        source_ref = candidate.source_ref
        order_qty = candidate.source_data.get("order_qty")
        material_available = candidate.source_data.get("material_available")

        if order_qty == material_available:
            return JobResult(is_success=False, error_message="no mismatch left to fix")

        rpatool_payload = {
            "source_ref": str(source_ref),
            "target_order_qty": material_available,
        }

        return JobResult(is_success=True, rpatool_payload=rpatool_payload)
    

    def verify_result(self, handover_job: HandoverJob) -> JobResult:
    
        job_id = handover_job.job_id

        # get erp order number/id
        rpatool_payload = handover_job.rpatool_payload
        if not rpatool_payload:
            return JobResult(is_success=False, error_message="missing rpatool_payload")
        
        # get the order number/id and the target qty sent to RPA tool
        source_ref = rpatool_payload.get("source_ref")
        target_order_qty = rpatool_payload.get("target_order_qty")


        # get the 'real' qty now in erp
        order_qty_erp = self.erp_backend.get_order_qty(source_ref)

        if order_qty_erp is None:
            return JobResult(
                is_success=False, error_code="VERIFICATION_MISMATCH", error_message=f"Order {source_ref} not found in ERP")

        # compare them
        if order_qty_erp != target_order_qty:
            error_message= f"ERP shows mismatch. {source_ref} should be {target_order_qty}, is {order_qty_erp}"
            self.logger.system(error_message, job_id)
            return JobResult(is_success=False, error_code="VERIFICATION_MISMATCH", error_message=error_message)

        self.logger.system(f"OK. Should be: {target_order_qty}, is: {order_qty_erp}", job_id)
        return JobResult(is_success=True)


# ============================================================
# HANDOVER
# ============================================================

class HandoverRepository:
    """Persist and validate the file-based state shared with the RPA tool."""
    HANDOVER_FILE = "handover.json"

    def __init__(self, logger) -> None:
        self.logger = logger


    def read(self) -> HandoverJob:
        ''' read HANDOVER_FILE '''
        
        last_err=None

        for attempt in range(7):
            try:
                # read file
                with open(self.HANDOVER_FILE, "r", encoding="utf-8") as f:
                    handover_data = json.load(f)
                
                # rebuild object
                handover_job = self._validate_and_build_handover_job(handover_data)

                return handover_job
                
            except Exception as err:
                last_err = err
                self.logger.system(f"WARN: retry {attempt+1}/7 : {err}")
                time.sleep(attempt/10)
        
        
        raise RuntimeError(f"{self.HANDOVER_FILE} unreadable: {last_err}")
    
      
    def write(self, handover_job: HandoverJob) -> None:
        ''' atomic write of HANDOVER_FILE '''

        handover_data = asdict(handover_job)

        self._validate_and_build_handover_job(handover_data) # only validate (ignore return)
        job_id = handover_data.get("job_id")

        last_err = None
        
        for attempt in range(7):
            temp_path = None
            try:
                
                dir_path = os.path.dirname(os.path.abspath(self.HANDOVER_FILE))
                fd, temp_path = tempfile.mkstemp(dir=dir_path, suffix=".tmp")

                #atomic write
                with os.fdopen(fd, "w", encoding="utf-8") as tmp:
                    json.dump(handover_data, tmp, indent=2) # indent for human eyes
                    tmp.flush()
                    os.fsync(tmp.fileno())

                os.replace(temp_path, self.HANDOVER_FILE)
                
                self.logger.system(
                    f"wrote handover state={handover_data.get('state')} for job_type={handover_data.get('job_type')} "
                    f"with rpatool_payload={handover_data.get('rpatool_payload')}",
                    job_id,
                )               
                return

            except Exception as err:
                last_err = err
                self.logger.system(f"WARN: {attempt+1}/7 error", job_id)
                time.sleep(attempt/10) # 0 0.1... 0.6 sec     

            finally:
                if temp_path and os.path.exists(temp_path):
                    try: os.remove(temp_path)
                    except Exception: pass

        self.logger.system(f"CRITICAL: cannot write {self.HANDOVER_FILE} {last_err}", job_id)
        raise RuntimeError(f"CRITICAL: cannot write {self.HANDOVER_FILE}")


    def _validate_and_build_handover_job(self, handover_data: dict) -> HandoverJob:
        """Validate raw handover dict and return HandoverJob."""

        state = handover_data.get("state")
        job_id = handover_data.get("job_id")
        job_type = handover_data.get("job_type")
        job_source_type = handover_data.get("job_source_type")
        source_ref = handover_data.get("source_ref")
        email_address = handover_data.get("email_address")
        email_subject = handover_data.get("email_subject")
        email_body = handover_data.get("email_body")
        source_data = handover_data.get("source_data")
        rpatool_payload = handover_data.get("rpatool_payload")

        if state is None:
            raise ValueError("state missing")

        if state not in get_args(HandoverState):
            raise ValueError(f"unknown state: {state}")

        if job_id is not None:
            try:
                job_id = int(job_id)
            except Exception:
                raise ValueError(f"job_id not INT-like: {job_id}")

        if state == "idle":
            if any(v is not None for v in (
                job_id, job_type, job_source_type, source_ref,
                email_address, email_subject, email_body, source_data, rpatool_payload
            )):
                raise ValueError(f"state 'idle' should have no more variables: {handover_data}")

        elif state in ("job_queued", "job_running", "job_verifying"):
            required_fields = {
                "job_id": job_id,
                "job_type": job_type,
                "job_source_type": job_source_type,
                "source_ref": source_ref,
                "source_data": source_data,
                "rpatool_payload": rpatool_payload,
            }

            missing = [k for k, v in required_fields.items() if v is None]
            if missing:
                raise ValueError(f"{state} has missing fields in {self.HANDOVER_FILE}: {missing}")

            if job_type not in get_args(JobType):
                raise ValueError(f"unknown job_type: {job_type}")

            if job_source_type not in get_args(JobSourceType):
                raise ValueError(f"unknown job_source_type: {job_source_type}")

            if job_source_type in ("personal_inbox", "shared_inbox"):
                required_fields = {
                    "email_address": email_address,
                    "email_subject": email_subject,
                    "email_body": email_body,
                }

                missing = [k for k, v in required_fields.items() if v is None]
                if missing:
                    raise ValueError(f"{job_source_type} has missing fields in {self.HANDOVER_FILE}: {missing}")
                
            if not isinstance(source_data, dict):
                raise ValueError("source_data must be dict")
            if not isinstance(rpatool_payload, dict):
                raise ValueError("rpatool_payload must be dict")

        return HandoverJob(
            state=state,
            job_id=job_id,
            job_type=job_type,
            job_source_type=job_source_type,
            source_ref=source_ref,
            email_address=email_address,
            email_subject=email_subject,
            email_body=email_body,
            source_data=source_data,
            rpatool_payload=rpatool_payload,
        )


    def is_valid_transition(self, prev_state: HandoverState | None, state: HandoverState) -> bool:
        """ transition-validator for RobotRuntime loop. Only runs when state != prev_state. """

        if prev_state is None: # at startup
            return True

        allowed_transitions: dict[HandoverState, set[HandoverState]] = {
            "idle": {"job_queued", "safestop"},
            "job_queued": {"job_running", "safestop"},
            "job_running": {"job_verifying", "safestop"},
            "job_verifying": {"idle", "safestop"},
            "safestop": {"idle"},
        }

        allowed_next = allowed_transitions[prev_state]

        if state not in allowed_next:
            return False
        
        return True


# ============================================================
# USER NOTIFICATIONS
# ============================================================

class UserNotificationService:
    """Only for personal_inbox user-facing replies."""

    ADMIN_EMAIL = "admin_rpa@company.local"
    COMMAND_JOB_ID = 999999999999

    def __init__(self, mail_backend_personal, recordings_destination_folder: str, rpa_tool_execution_timeout: int):
        self.mail_backend_personal = mail_backend_personal
        self.recordings_destination_folder = recordings_destination_folder
        self.rpa_tool_execution_timeout = rpa_tool_execution_timeout


    def send_final_reply_and_delete_original(self, candidate: JobCandidate, job_status: JobStatus, error_code: str | None, job_id: int, reason=None, from_safestop:bool=False, from_initialize:bool=False, delete_after=True) -> None:
        
        subject, body = self._build_job_reply(
            job_status=job_status,
            error_code=error_code,
            job_id=job_id,
            reason=reason,
            from_safestop=from_safestop,
            from_initialize=from_initialize,
        )

        self._send(candidate, subject, body, job_id, delete_after)


    def send_recovery_reply(self, audit_row: dict, candidate: JobCandidate, from_safestop:bool, from_initialize:bool, delete_after:bool) -> None:

        self.send_final_reply_and_delete_original(
            candidate=candidate,
            job_status=audit_row["job_status"],
            error_code=audit_row.get("error_code"),
            job_id=audit_row["job_id"],
            reason=audit_row.get("error_message"),
            from_safestop=from_safestop,
            from_initialize=from_initialize,
            delete_after=delete_after,
        )


    def send_out_of_service_reply(self, candidate: JobCandidate, job_id: int) -> None:
        
        self.send_final_reply_and_delete_original(
            candidate=candidate,
            job_status="FAIL",
            error_code="OUT_OF_SERVICE",
            job_id=job_id,
            )


    def send_command_reply(self, candidate: JobCandidate) -> None:
        self._send(
            candidate=candidate,
            subject="got it!",
            body="Command received.",
            job_id=self.COMMAND_JOB_ID,
            delete_after=True,
        )


    def send_admin_alert(self, reason: str) -> None:
        fake_candidate = JobCandidate(
            source_ref="safestop, no real source_ref",
            email_address=self.ADMIN_EMAIL,
            email_subject="",
            email_body="",
            job_source_type="personal_inbox",
            source_data={},
        )

        body = (
            "Robot is in degraded mode.\n\n"
            f"Reason:\n{reason}\n\n"
            "Available commands: 'stop1234' and 'restart1234'."
        )

        self._send(
            candidate=fake_candidate,
            subject="safestop notice",
            body=body,
            job_id=self.COMMAND_JOB_ID,
            delete_after=False,
        )


    def send_online_notice(self, candidate: JobCandidate, job_id: int) -> None:
        body = (
            ">Hello, human<\n\n"
            "The first request each day is replied with: online\n"
            "Next message is sent after completion\n"
            f"(in max {self.rpa_tool_execution_timeout} seconds from now).\n"
        )

        self._send(
            candidate=candidate,
            subject="ONLINE",
            body=body,
            job_id=job_id,
            delete_after=False,
        )


    def _classify_reply_kind(self, job_status:JobStatus, error_code:str|None) -> str:
        if job_status == "DONE":
            return "DONE"
        
        if job_status == "FAIL" and error_code == "PRE_HANDOVER_CRASH":
            return "NOT_STARTED"

        if job_status == "FAIL" and error_code == "OUT_OF_SERVICE":
            return "OUT_OF_SERVICE"

        if job_status == "FAIL" and error_code == "RPA_TOOL_CRASH":
            return "STARTED_BUT_CRASHED"

        if job_status == "FAIL" and error_code == "VERIFICATION_MISMATCH":
            return "VERIFICATION_MISMATCH"

        if job_status == "FAIL" and error_code == "POST_HANDOVER_CRASH":
            return "VERIFYING_CRASH"

        if job_status == "REJECTED":
            return "NOT_STARTED"

        if job_status == "QUEUED":
            return "NOT_STARTED"

        if job_status == "RUNNING":
            return "STARTED_BUT_CRASHED"

        if job_status == "VERIFYING":
            return "VERIFYING_CRASH"

        if job_status == "FAIL":
            return "UNKNOWN_FAIL"

        raise ValueError(f"Cannot classify reply for job_status={job_status}, error_code={error_code}")


    def _build_job_reply(self, job_status: JobStatus, error_code: str | None, job_id: int, reason, from_safestop:bool, from_initialize:bool) -> tuple[str, str]:
        # TODO: for increased user value, extend reply with a short summary eg. "changed PO 450221 on SKU 110212 from 34pcs to 31pcs"
        
        subject: str
        body: str

        recording_text = self._get_recording_text(job_id)
        reply_kind = self._classify_reply_kind(job_status, error_code)

        if reply_kind == "DONE":
            subject = "DONE"
            body = (
                    f"Job completed successfully.\n\n"
                    f"Job ID: {job_id}\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )


        elif reply_kind == "NOT_STARTED":
            subject = "FAIL"
            body = (
                    f"Your request was not started.\n\n"
                    f"{f'Reason: {reason}\n\n' if reason else ''}"
                    f"Job ID: {job_id}\n"
                    f"Keep calm, no changes were made in ERP.\n\n"
                    f"This email can be deleted."
                )


        elif reply_kind == "STARTED_BUT_CRASHED":
            subject = "FAIL"
            body = (
                    f"The robot started your request, but then crashed.\n\n"
                    f"{f'Reason: {reason}\n\n' if reason else ''}"
                    f"Job ID: {job_id}\n"
                    f"Changes may have been made in ERP before the crash.\n"
                    f"It is (very) recommended that you review the result manually.\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )

                
        elif reply_kind == "VERIFICATION_MISMATCH":
            subject = "FAIL"
            body = (
                    f"The robot completed the request, and the result was checked in ERP.\n"
                    f"However, the final ERP data did not match the expected result.\n\n"
                    f"{f'Reason: {reason}\n\n' if reason else ''}"
                    f"Job ID: {job_id}\n"
                    f"You NEED TO review the result manually in ERP.\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )

        
        elif reply_kind == "VERIFYING_CRASH":
            subject = "FAIL"
            body = (
                    f"The robot completed the request, but crashed during the final verification stage.\n"
                    f"The outcome could therefore not be confirmed automatically.\n\n"
                    f"Job ID: {job_id}\n"
                    f"Please verify the result manually in ERP.\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )


        elif reply_kind == "OUT_OF_SERVICE":
            subject = "FAIL"
            body = (
                    "The robot is out-of-service and does not accept any new jobs.\n\n"
                    "This email can be deleted."
                )

        elif reply_kind == "UNKNOWN_FAIL":
            subject = "FAIL"
            body = (
                    f"The robot crashed and the exact job outcome could not be classified.\n\n"
                    f"{f'Reason: {reason}\n\n' if reason else ''}"
                    f"Job ID: {job_id}\n"
                    f"Please review the result manually in ERP.\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )

        else:
            raise ValueError(f"Unhandled reply_kind={reply_kind}")

        if from_safestop:
            body = body.replace(
                    "This email can be deleted.",

                    "To avoid further problems, the robot will go out-of-service.\n"
                    "This email can be deleted.",
                )

        if from_initialize:
            body = (
                    "The Robot was offline and has now restarted.\n"
                    "If you already received a final reply (DONE/FAIL) for this job, you can ignore this recovery message."
                ) + "\n\n" + body

        return subject, body

    def _get_recording_text(self, job_id: int) -> str:
        recording_path = self._get_recording_path(job_id)
        if not recording_path:
            return ""

        return (
            "A screen recording is available for review:\n"
            f"{recording_path}\n\n"
        )

    def _get_recording_path(self, job_id: int) -> str | None:
        path = Path(self.recordings_destination_folder) / f"{job_id}.mp4"
        if path.exists():
            return str(path)
        return None

    def _send(self, candidate: JobCandidate, subject: str, body: str, job_id: int, delete_after: bool,) -> None:
        if delete_after:
            self.mail_backend_personal.reply_and_delete(
                candidate=candidate,
                extra_subject=subject,
                extra_body=body,
                job_id=job_id,
            )
        else:
            self.mail_backend_personal.send_reply(
                candidate=candidate,
                extra_subject=subject,
                extra_body=body,
                job_id=job_id,
            )


# ============================================================
# RECORDING / SAFESTOP / INFRASTRUCTURE
# ============================================================   
                      
class RecordingService:
    ''' screen-recording to capture all RPA tool screen-activity '''

    RECORDINGS_IN_PROGRESS_FOLDER = "recordings_in_progress"
    RECORDINGS_DESTINATION_FOLDER = "recordings_destination"

    def __init__(self, logger,) -> None:
        #written by AI

        self.logger = logger
        self.recording_process = None
        self._ffmpeg_warned = False


    def _get_screen_resolution(self):
        try:
            output = subprocess.check_output(["xrandr"], text=True)
            for line in output.splitlines():
                if "*" in line:
                    res = line.split()[0]
                    return res.split("x")
        except Exception:
            pass

        # fallback: Tkinter
        try:
            root = tk.Tk()
            root.withdraw()
            width = root.winfo_screenwidth()
            height = root.winfo_screenheight()
            root.destroy()
            return str(width), str(height)
        except Exception:
            pass

        return "1920", "1080"

 
    def start(self, job_id) -> None:
        """start the screen recording"""
        # written by AI

        os.makedirs(self.RECORDINGS_IN_PROGRESS_FOLDER, exist_ok=True)
        filename = f"{self.RECORDINGS_IN_PROGRESS_FOLDER}/{job_id}.mp4"

        drawtext = (
            f"drawtext=text='job_id  {job_id}':"
            "x=200:y=20:"
            "fontsize=32:"
            "fontcolor=lightyellow:"
            "box=1:"
            "boxcolor=black@0.5"
        )

        if platform.system() == "Windows":
            ffmpeg_path = None

            local_ffmpeg = Path("./ffmpeg.exe")
            if local_ffmpeg.exists():
                ffmpeg_path = str(local_ffmpeg)
            else:
                ffmpeg_in_path = shutil.which("ffmpeg")
                if ffmpeg_in_path:
                    ffmpeg_path = ffmpeg_in_path

            if ffmpeg_path is None:
                if not self._ffmpeg_warned:
                    message = (
                        "FFMPEG NOT FOUND\n\n"
                        "Screen recording is disabled.\n\n"
                        "Fix:\n"
                        "1. Go to: https://www.gyan.dev/ffmpeg/builds/\n"
                        "2. Download 'ffmpeg-git-essentials'\n"
                        "3. Extract the archive\n"
                        "4. Open the 'bin' folder\n"
                        "5. Copy ffmpeg.exe next to main.py\n"
                    )

                    print("\n" + "="*60 + "\n" + message + "\n" + "="*60 + "\n")
                    self.logger.system(message, job_id)
                    self.logger.ui("--> recording disabled (ffmpeg missing)")
                    self._ffmpeg_warned = True
                return

            capture = ["-f", "gdigrab", "-i", "desktop"]

            recording_process = subprocess.Popen(
                [
                    ffmpeg_path,
                    "-y",
                    *capture,
                    "-framerate", "15",
                    "-vf", drawtext,
                    "-vcodec", "libx264",
                    "-pix_fmt", "yuv420p",
                    "-preset", "ultrafast",
                    filename,
                ],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
            )

        else:
            display = os.environ.get("DISPLAY")
            if not display:
                self.logger.system("WARN: screen-recording disabled because DISPLAY is missing", job_id)
                return

            ffmpeg_path = shutil.which("ffmpeg")
            if ffmpeg_path is None:
                self.logger.system("WARN: screen-recording disabled because ffmpeg is not installed", job_id)
                return

            width, height = self._get_screen_resolution()

            capture = [
                "-video_size", f"{width}x{height}",
                "-f", "x11grab",
                "-i", display,
            ]

            recording_process = subprocess.Popen(
                [
                    ffmpeg_path,
                    "-y",
                    *capture,
                    "-framerate", "15",
                    "-vf", drawtext,
                    "-vcodec", "libx264",
                    "-preset", "ultrafast",
                    filename,
                ],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                start_new_session=True,
            )

        time.sleep(0.2)

        if recording_process.poll() is not None:
            self.logger.system("WARN: ffmpeg exited immediately; recording did not start", job_id)
            return

        self.recording_process = recording_process
        self.logger.system("recording started", job_id)

        
    def stop(self, job_id=None) -> None:
        ''' allow global kill of FFMPEG processes since Orchestrator is designed to run on a dedicated machine '''
        # written by AI

        try: self.logger.system("stop recording", job_id)
        except Exception: pass

        recording_process = self.recording_process
        self.recording_process = None

        try:
            if recording_process is not None:
                # try first stop only our own process
                if platform.system() == "Windows":
                    try:
                        recording_process.send_signal(
                            getattr(signal, "CTRL_BREAK_EVENT", signal.SIGTERM)
                        )
                    except Exception:
                        try:
                            recording_process.terminate()
                        except Exception:
                            pass

                    try:
                        recording_process.wait(timeout=8)
                        return
                    except subprocess.TimeoutExpired:
                        pass

                    # else, kill only our own process
                    try:
                        subprocess.run(
                            ["taskkill", "/PID", str(recording_process.pid), "/T", "/F"],
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                            check=False,
                        )
                        recording_process.wait(timeout=3)
                        return
                    except Exception:
                        pass

                    # last resort, global kill all ffmpeg
                    subprocess.run(
                        ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )

                else:
                    # try first stop only our own process
                    try:
                        os.killpg(recording_process.pid, signal.SIGINT)

                    except Exception:
                        try:
                            recording_process.terminate()
                        except Exception:
                            pass

                    try:
                        recording_process.wait(timeout=8)
                        return
                    except subprocess.TimeoutExpired:
                        pass

                    # else, kill only our own process
                    try:
                        os.killpg(recording_process.pid, signal.SIGKILL)
                        recording_process.wait(timeout=3)
                        return
                    except Exception:
                        pass

                    # last resort, global kill all ffmpeg
                    subprocess.run(
                        ["killall", "-q", "-KILL", "ffmpeg"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )

            else:
                # fallback if process object is lost
                if platform.system() == "Windows":
                    subprocess.run(
                        ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )
                else:
                    subprocess.run(
                        ["killall", "-q", "-KILL", "ffmpeg"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )
        except Exception as err:
            self.logger.system(f"WARN from stop(): {err}", job_id)


    def upload_recording(self, job_id, max_attempts=3) -> None:
        ''' upload to a shared drive'''
    
        local_file = f"{self.RECORDINGS_IN_PROGRESS_FOLDER}/{job_id}.mp4"
        local_file = Path(local_file)

        remote_path = Path(self.RECORDINGS_DESTINATION_FOLDER) / f"{job_id}.mp4"
        remote_path.parent.mkdir(parents=True, exist_ok=True)

        for attempt in range(max_attempts):
            try:
                
                shutil.copy2(local_file, remote_path)
                self.logger.system(f"upload success: {remote_path}", job_id)
                try: os.remove(local_file)
                except Exception: pass

                return

            except Exception as e:
                self.logger.system(f"Attempt {attempt+1}/{max_attempts} failed: {e}", job_id)
                time.sleep(attempt + 1)
        
        self.logger.system(f"upload failed: {remote_path}", job_id)


    def cleanup_aborted_recordings(self):
        """Upload or clean up recordings left behind by aborted runs."""

        directory = Path(self.RECORDINGS_IN_PROGRESS_FOLDER)
        if not directory.exists():
            return
        
        for file in directory.iterdir():

            if file.is_file() and file.suffix == ".mp4":
                job_id = file.stem
                
                try:
                    self.logger.system(f"cleanup upload started")
                    self.upload_recording(job_id)
                except Exception as err:
                    self.logger.system(f"cleanup failed for {job_id}: {err}")


class FriendsRepository:
    '''Example access-control source for personal_inbox'''

    def __init__(self) -> None:
        self.access_by_email: dict[str, set[str]] = {}
        self.access_file_mtime: float | None = None


    def _ensure_friends_file_exists(self, path: str = "friends.xlsx") -> None:
        '''Create a template friends.xlsx if missing.'''
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws["A1"] = "email"
        ws["B1"] = "ping"
        ws["C1"] = "job1"
        ws["D1"] = "job2"

        ws["A2"] = "alice@example.com"
        ws["B2"] = "x"

        ws["A3"] = "bob@test.com"
        ws["B3"] = "x"
        ws["C3"] = "x"
        ws["D3"] = "x"

        wb.save(path)
        wb.close()


    def _load_access_file(self, filepath: str) -> dict[str, set[str]]:
        '''
        Reads friends.xlsx and returns for example:

        {
            "alice@example.com": {"ping"},
            "ex2@whatever.com": {"ping", "job1"}
        }
        '''
        # written by AI

        wb = load_workbook(filepath, data_only=True)
        try:
            ws = wb.active
            assert ws is not None

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                raise ValueError("friends.xlsx contains no users")

            header = rows[0]
            self._validate_friends_header(header)
            access_map: dict[str, set[str]] = {}

            for row in rows[1:]:
                email_cell = row[0]
                if email_cell is None:
                    continue

                email = str(email_cell).strip().lower()
                if not email:
                    continue

                permissions: set[str] = set()

                for col in range(1, len(header)):
                    jobname = header[col]
                    if jobname is None:
                        continue

                    jobname = str(jobname).strip().lower()
                    cell = row[col] if col < len(row) else None

                    if cell is None:
                        continue

                    if str(cell).strip().lower() == "x":
                        permissions.add(jobname)

                access_map[email] = permissions

            return access_map
        finally:
            wb.close()


    def reload_if_modified(self) -> bool:
        '''Reload friends.xlsx if changed.'''
        # written by AI

        path = "friends.xlsx"
        self._ensure_friends_file_exists(path)

        mtime = os.path.getmtime(path)
        if self.access_file_mtime == mtime:
            return False

        new_access = self._load_access_file(path)
        self._validate_friends_access(new_access)

        self.access_by_email = new_access
        self.access_file_mtime = mtime

        return True


    def is_allowed_sender(self, email_address: str | None) -> bool:

        if not email_address:
            return False
        
        email = email_address.strip().lower()        
        return email in self.access_by_email


    def has_job_access(self, email_address: str, job_type: str) -> bool:
        email = email_address.strip().lower()
        job = job_type.strip().lower()
        return job in self.access_by_email.get(email, set())


    def _validate_friends_access(self, access_map: dict[str, set[str]]) -> None:
        ''' not implemented '''
        if not isinstance(access_map, dict):
            raise ValueError("access_map must be dict")

        valid_job_types = set(get_args(JobType))

        for email, permissions in access_map.items():
            if not isinstance(email, str):
                raise ValueError(f"invalid email key type: {email}")

            email_normalized = email.strip().lower()
            if not email_normalized:
                raise ValueError("empty email in access_map")

            if "@" not in email_normalized:
                raise ValueError(f"invalid email in friends.xlsx: {email}")

            if not isinstance(permissions, set):
                raise ValueError(f"permissions must be set for {email}")

            invalid_permissions = permissions - valid_job_types
            if invalid_permissions:
                raise ValueError(
                    f"invalid job types for {email}: {sorted(invalid_permissions)}. "
                    f"Allowed: {sorted(valid_job_types)}"
                )
            

    def _validate_friends_header(self, header_row) -> None:
        ''' validate headers '''
        if not header_row or str(header_row[0]).strip().lower() != "email":
            raise ValueError("friends.xlsx column A must be 'email'")

        valid_job_types = set(get_args(JobType))

        for col in range(1, len(header_row)):
            jobname = header_row[col]
            if jobname is None:
                continue

            jobname_str = str(jobname).strip().lower()
            if jobname_str not in valid_job_types:
                raise ValueError(
                    f"invalid job type column in friends.xlsx: {jobname_str}. "
                    f"Allowed: {sorted(valid_job_types)}"
                )


class NetworkService:
    """Check whether the machine currently has access to the required company network resources."""
    # Placeholder for implementation
    
    # e.g. NETWORK_HEALTHCHECK_PATH=    r"G:\\"    or    r"\\\\server\\share"
    NETWORK_HEALTHCHECK_PATH = None

    def __init__(self, logger) -> None:
        self.logger = logger
        self.network_state = False #assume offline at start
        self.next_network_check_time = 0


    def has_network_access(self) -> bool:
        #this runs at highest once every hour (if online), or before new jobs

        now = time.time()

        if now < self.next_network_check_time:
            return self.network_state

        try:
            if self.NETWORK_HEALTHCHECK_PATH is None: # demo assumption
                online = True                         # demo assumption
            else:
                os.listdir(self.NETWORK_HEALTHCHECK_PATH)
                online = True 

        except Exception:
            online = False
            
        # update log if any network change (and UI? )
        if online != self.network_state:
            self.network_state = online

            if online:
                self.logger.system("network restored")
            else:
                self.logger.system(f"WARN: network lost")

        # check every minute if offline, else every hour
        if online:
            self.next_network_check_time = now + 3600   # 1 h
        else:
            self.next_network_check_time = now + 60     # 1 min
        
        return online


class AuditRepository:
    ''' handles job_audit.db, an audit-style activity log '''

    DB_PATH = "job_audit.db"

    def __init__(self, logger) -> None:
        self.logger = logger


    def _connect_with_retry(self) -> sqlite3.Connection:
  
        max_retries = 3
        for attempt in range(max_retries):
            try:
                conn = sqlite3.connect(self.DB_PATH, timeout=10)
                return conn
            except sqlite3.OperationalError as e:
                if attempt == max_retries - 1:
                    raise
                time.sleep(0.5)

        conn = sqlite3.connect(self.DB_PATH, timeout=10,)
        return conn  
        

    def ensure_db_exists(self) -> None:
        
        with self._connect_with_retry() as conn:
            cur = conn.cursor()
           
            cur.execute('''
                CREATE TABLE IF NOT EXISTS audit_log
                         (
                        job_id INTEGER PRIMARY KEY, 
                        job_type TEXT, 
                        job_status TEXT, 
                        email_address TEXT, 
                        email_subject TEXT, 
                        source_ref TEXT,
                        job_start_date TEXT, 
                        job_start_time TEXT, 
                        job_finish_time TEXT, 
                        final_reply_sent INTEGER NOT NULL DEFAULT 0,
                        job_source_type TEXT,
                        error_code TEXT, 
                        error_message TEXT 
                        )
                        ''')


    def _build_audit_fields(self, job_id, email_address=None, email_subject=None, source_ref=None, job_type: JobType | None = None, job_start_date=None, job_start_time=None, job_finish_time=None, job_status: JobStatus | None = None, final_reply_sent=None, job_source_type: JobSourceType | None = None, error_code=None, error_message=None,) -> dict:
        all_fields = {
            "job_id": job_id,
            "email_address": email_address,
            "email_subject": email_subject,
            "source_ref": source_ref,
            "job_type": job_type,
            "job_start_date": job_start_date,
            "job_start_time": job_start_time,
            "job_finish_time": job_finish_time,
            "job_status": job_status,
            "final_reply_sent": final_reply_sent,
            "job_source_type": job_source_type,
            "error_code": error_code,
            "error_message": error_message,
        }

        # drop None:s
        fields = {k: v for k, v in all_fields.items() if v is not None}

        self.logger.system(f"received audit fields {fields}", job_id) # TODO: ensure email_address och email_subject are GDPR safe 

        return fields


    def insert_job(self, job_id, email_address=None, email_subject=None, source_ref=None, job_type: JobType | None=None, job_start_date=None, job_start_time=None, job_finish_time=None, job_status: JobStatus | None=None, final_reply_sent=None, job_source_type:JobSourceType | None=None, error_code=None, error_message=None,) -> None:
        # use for new row

        fields = self._build_audit_fields(
            job_id=job_id,
            email_address=email_address,
            email_subject=email_subject,
            source_ref=source_ref,
            job_type=job_type,
            job_start_date=job_start_date,
            job_start_time=job_start_time,
            job_finish_time=job_finish_time,
            job_status=job_status,
            final_reply_sent=final_reply_sent,
            job_source_type=job_source_type,
            error_code=error_code,
            error_message=error_message,
        )
        
        columns = ", ".join(fields.keys())
        placeholders = ", ".join("?" for _ in fields)

        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                f"INSERT INTO audit_log ({columns}) VALUES ({placeholders})",
                tuple(fields.values())
            )


    def update_job(self, job_id, email_address=None, email_subject=None, source_ref=None, job_type: JobType | None=None, job_start_date=None, job_start_time=None, job_finish_time=None, job_status: JobStatus | None=None, final_reply_sent=None, job_source_type:JobSourceType | None=None, error_code=None, error_message=None,) -> None:
        # example use: self.audit_repo.update_job(job_id=20260311124501, job_type="job1")

        fields = self._build_audit_fields(
            job_id=job_id,
            email_address=email_address,
            email_subject=email_subject,
            source_ref=source_ref,
            job_type=job_type,
            job_start_date=job_start_date,
            job_start_time=job_start_time,
            job_finish_time=job_finish_time,
            job_status=job_status,
            final_reply_sent=final_reply_sent,
            job_source_type=job_source_type,
            error_code=error_code,
            error_message=error_message,
        )
        
        fields.pop("job_id", None)

        if not fields:
            return

        set_clause = ", ".join(f"{k}=?" for k in fields)

        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                f"UPDATE audit_log SET {set_clause} WHERE job_id=?",
                (*fields.values(), job_id)
            )

            if cur.rowcount == 0:
                raise ValueError(f"update_job(): no row in DB with job_id={job_id}")


    def count_done_jobs_today(self) -> int:
        today = datetime.date.today().isoformat()

        with self._connect_with_retry() as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT COUNT(*)
                FROM audit_log
                WHERE job_start_date = ?
                AND job_status = 'DONE'
            ''', (today,))
            
            result = cur.fetchone()[0]

        return result


    def has_sender_job_today(self, sender_mail, job_id) -> bool:

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE job_start_date = ? AND email_address = ? AND job_id != ?
                ''',
                (today, sender_mail, job_id,)
            )

            jobs_today = cur.fetchone()[0]

        return jobs_today > 0


    def has_been_processed_today(self, source_ref) -> bool:
        # use to avoid bad loops in query-jobs

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE job_start_date = ? AND source_ref = ?
                ''',
                (today, source_ref,)
            )

            jobs_today = cur.fetchone()[0]

        return jobs_today > 0


    def get_latest_job_id(self) -> int:
        with self._connect_with_retry() as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT job_id
                FROM audit_log
                ORDER BY job_id DESC
                LIMIT 1
            ''')
            row = cur.fetchone()

        return row[0] if row is not None else 0


    def get_pending_reply_jobs(self) -> list[dict]:
        job_source_type: JobSourceType = "personal_inbox"

        with self._connect_with_retry() as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                '''
                SELECT job_id, job_source_type, email_address, email_subject, source_ref, job_status, error_code, error_message
                FROM audit_log
                WHERE job_source_type = ?
                AND COALESCE(final_reply_sent, 0) = 0
                ORDER BY job_id
                ''',
                (job_source_type,)
            )
            rows = cur.fetchall()

        list_of_dicts = [dict(row) for row in rows]

        return list_of_dicts


class LoggerService:
    """ logging functions"""
    def __init__(self, dashboard_ui) -> None:
        self.dashboard_ui = dashboard_ui


    def ui(self, text:str, blank_line_before: bool = False) -> None:
        
        self.dashboard_ui.post_log_line(text, blank_line_before)


    def system(self, event_text, job_id: int | None=None,):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        event_text = str(event_text)

        # get caller function name
        try:
            frame = sys._getframe(1)
            caller_name = frame.f_code.co_name
            instance = frame.f_locals.get("self")
            if instance is not None:
                class_name = instance.__class__.__name__
                caller = f"{class_name}.{caller_name}()"
            else:
                caller = f"{caller_name}()"

        except Exception:
            caller = "unknown_caller()"
      
        log_line = f"{timestamp} | py  | job_id={job_id or ''} | {caller} | {event_text}"

        # normalize to single-line
        log_line = " ".join(str(log_line).split())

        last_err = None
        for i in range(7):
            try:
                with open("system.log", "a", encoding="utf-8") as f:
                    f.write(log_line + "\n")
                    f.flush()
                return

            except Exception as err:
                last_err = err
                print(f"WARN: retry {i+1}/7 from log_system():", err)
                time.sleep(i + 1)

        # fallback to print() when log fails        
        print(f"[print fallback] {job_id} {event_text} | {last_err}")  
 

class SafeStopController:
    """Handle degraded mode, crash recovery, and operator restart/stop commands."""
    def __init__(self, logger, recording_service, hide_recording_overlay, post_status_update, set_ui_shutdown, mail_backend_personal, audit_repo, generate_job_id, friends_repo, notification_service, check_for_stop_flag) -> None:
        self.logger = logger
        self.recording_service = recording_service
        self._hide_recording_overlay = hide_recording_overlay
        self._post_status_update = post_status_update
        self._set_ui_shutdown = set_ui_shutdown
        self.mail_backend_personal = mail_backend_personal
        self.audit_repo = audit_repo
        self._generate_job_id = generate_job_id
        self.friends_repo = friends_repo
        self._degraded_mode_entered = False
        self.notification_service = notification_service
        self._check_for_stop_flag = check_for_stop_flag


    def run_degraded_mode(self, fault: RuntimeFault,) -> None:
        '''
        Rules:
        * no job intake
        * mail-flow inactivated
        * query-flow inactivated
        * no handover to RPA tool
        * 'safestop' status text in UI
        * STOP and RESTART commands allowed 
        * send rejected reply to email from users in friends.xlsx
        * warning email sent to admin
        '''
        
        if self._degraded_mode_entered: return
        self._degraded_mode_entered = True

        job_id = fault.job_id
        handover_job = fault.handover_job
        error_message = fault.error_message

        err_with_traceback_ext = f"RobotRuntime crashed:\n\nfault={fault}\n\n{fault.traceback_text}"
        
        if handover_job is not None and handover_job.state != "idle":
            err_with_traceback_ext += (
                f"\n\n...while working on job_type={handover_job.job_type} "
                f"with rpatool_payload=\n{handover_job.rpatool_payload}"
            )

            # stop RPA_tool from possibly claiming the workflow. Intentionally not using handover_repo in degraded mode
            if handover_job.state == "job_queued":
                try:
                    handover_job.state="safestop"
                    handover_data = asdict(handover_job)

                    with open("handover.json.tmp", "w", encoding="utf-8") as f:
                        json.dump(handover_data, f, indent=2)

                    os.replace("handover.json.tmp", "handover.json")
    
                except Exception:
                    try: os.remove("handover.json") # remove file on failure
                    except Exception as e: self.logger.system(e)
           
        self.logger.system(err_with_traceback_ext, job_id)

        try: self.recording_service.stop()
        except Exception as e: self.logger.system(e, job_id)

        try: self.recording_service.cleanup_aborted_recordings()
        except Exception as e: self.logger.system(e, job_id)

        try: self.notification_service.send_admin_alert(err_with_traceback_ext)
        except Exception as e: self.logger.system(e, job_id)

        try:
            self.logger.ui("--> CRASH! All automations are halted. Admin is notified.")
            self.logger.ui(f"--> Reason: {error_message}")
        except Exception as e: self.logger.system(e, job_id)

        try: self.recover_pending_mail_jobs(fault, from_safestop=True) # register stuck mails in audit
        except Exception as e: self.logger.system(e, job_id)

        # placeholder for recovery logic for post_handover crash/mismatch for query jobs

        try: self._hide_recording_overlay()
        except Exception as e: self.logger.system(e, job_id)

        try: self._post_status_update("safestop")
        except Exception as e:
            self.logger.system(e, job_id)
            try: self._set_ui_shutdown()
            except Exception as e2: 
                self.logger.system(e2, job_id)
                os._exit(1)
            time.sleep(3)
            os._exit(0)
        
        self._enter_degraded_loop()
    
    
    def _mark_faulted_pending_job_for_recovery(self, fault: RuntimeFault):
        # update audit row to FAIL for pending reply jobs
        job_id = fault.job_id
        error_code = fault.error_code
        error_message = fault.error_message

        if job_id is not None:
            for audit_row in self.audit_repo.get_pending_reply_jobs(): # filter out final_reply_sent (eg. DONE)
                if job_id == audit_row.get("job_id"):
                    try: self.audit_repo.update_job(job_id=job_id, job_status="FAIL", error_code=error_code, error_message=error_message)
                    except Exception as e: self.logger.system(e, job_id)

    
    def recover_pending_mail_jobs(self, fault:RuntimeFault|None=None, from_safestop: bool = False, from_initialize: bool = False) -> None:

        # 0. If degraded mode was triggered by an active job, mark that audit row as failed
        if fault is not None:
            self._mark_faulted_pending_job_for_recovery(fault)

        # 1. Get current pending reply jobs and build lookup by source_ref
        pending_jobs_before_scan = self.audit_repo.get_pending_reply_jobs()
        pending_source_refs = {row["source_ref"] for row in pending_jobs_before_scan if row.get("source_ref") is not None}

        # 2. Register stuck mails in processing that are not yet represented in audit
        paths = self.mail_backend_personal.list_processing_mail_paths()

        for processing_path in paths:
            candidate = self.mail_backend_personal.parse_mail_file(processing_path)

            # silent delete non friends
            if not self.friends_repo.is_allowed_sender(candidate.email_address):
                try:
                    self.logger.ui(f"email recovered from {candidate.email_address}", blank_line_before=True)
                    self.logger.ui("--> rejected (not in friends.xlsx)")
                except Exception:
                    pass

                self.mail_backend_personal.delete_from_processing(candidate)
                continue

            # Skip if already in audit as pending reply
            if candidate.source_ref in pending_source_refs:
                continue

            job_id = self._generate_job_id()
            self._try_insert_recovery_audit_row(job_id, candidate, final_reply_sent=False, reason="RECOVERY")

            try:
                self.logger.ui(f"email recovered from {candidate.email_address}", blank_line_before=True)
                self.logger.ui("--> rejected (recovery)")
            except Exception:
                pass

            self.logger.system(f"email recovered from {candidate.email_address}", job_id)

        # 3. Re-read pending reply jobs, now including newly registered recovery rows
        pending_jobs_after_scan = self.audit_repo.get_pending_reply_jobs()

        for audit_row in pending_jobs_after_scan:
            job_id = audit_row.get("job_id")
            source_ref = audit_row.get("source_ref")

            path = Path(source_ref)
            if path.exists():
                candidate: JobCandidate
                candidate = self.mail_backend_personal.parse_mail_file(str(path))
                self.logger.system(f"re-built candidate={candidate.source_ref, candidate.email_address, candidate.email_subject} from actual email", job_id,)
                delete_after = True
            else:
                audit_row["error_code"] = "RECOVERY_SOURCE_MISSING"
                self.audit_repo.update_job(job_id=job_id, error_code="RECOVERY_SOURCE_MISSING")
                candidate = self._build_candidate_from_audit(audit_row)
                self.logger.system(f"re-built candidate={candidate} from audit due to missing processing file {source_ref}", job_id,)
                delete_after = False

            try:
                self.notification_service.send_recovery_reply(
                    audit_row,
                    candidate,
                    from_safestop,
                    from_initialize,
                    delete_after,
                    )
                self.logger.system("recovery reply sent", job_id)
                
                self.audit_repo.update_job(
                    job_id=job_id,
                    final_reply_sent=True,
                )
            except Exception as err:
                self.logger.system(f"recovery reply failed: {err}", job_id)
    

    def _build_candidate_from_audit(self, audit_row) -> JobCandidate:
        
        return JobCandidate(
            source_ref = audit_row.get("source_ref"),
            job_source_type = audit_row.get("job_source_type"),
            email_address = audit_row.get("email_address"),
            email_subject = audit_row.get("email_subject"),
            email_body = "[ORIGINAL MESSAGE LOST]",
            source_data = {},
            )


    def _check_for_restart_flag(self,) -> None:
        restartflag = "restart.flag"

        if os.path.isfile(restartflag):
            try: os.remove(restartflag)
            except Exception: pass
            self.logger.system(f"restart-command received from {restartflag}")
            self._restart_application()


    def _check_for_restart_command(self, candidate: JobCandidate) -> None:
        if candidate.email_subject is None:
            return

        if "restart1234" in candidate.email_subject.strip().lower():
            self.logger.system(f"restart command received from {candidate.email_address}")
            try: self.notification_service.send_command_reply(candidate)
            except Exception: pass
            self._restart_application()


    def _check_for_stop_command(self, candidate: JobCandidate) -> None:
        if candidate.email_subject is None:
            return

        if "stop1234" in candidate.email_subject.strip().lower():
            self.logger.system(f"stop command received from {candidate.email_address}")
            try: self.notification_service.send_command_reply(candidate)
            except Exception: pass
            try: self._set_ui_shutdown()
            except Exception: os._exit(1)
            os._exit(0)


    def _try_send_out_of_service_reply(self, candidate: JobCandidate, job_id: int) -> bool:
        final_reply_sent = False

        try:
            self.notification_service.send_out_of_service_reply(candidate, job_id)
            final_reply_sent = True

        except Exception as e:
            self.logger.system(e, job_id)
            
        return final_reply_sent


    def _try_insert_recovery_audit_row(self, job_id:int, candidate:JobCandidate, final_reply_sent: bool, reason,):
        
        if reason == "SAFESTOP":
            job_status="REJECTED"
            error_code="IN_SAFESTOP"
            error_message="not accepting new jobs in safestop"
        
        elif reason == "RECOVERY":
            job_status="FAIL"
            error_code="PRE_HANDOVER_CRASH"
            error_message="unknown, mail stuck in processing folder"
        
        else:
            raise ValueError(f"unknown reason: {reason}")


        try:
            now = datetime.datetime.now()
            job_source_type: JobSourceType = "personal_inbox" 
            
            self.audit_repo.insert_job(
                job_id=job_id,
                source_ref=candidate.source_ref,
                email_address=candidate.email_address,
                email_subject=candidate.email_subject,
                job_start_date=now.strftime("%Y-%m-%d"),
                job_start_time=now.strftime("%H:%M:%S"),
                job_status=job_status,
                error_code=error_code,
                error_message=error_message,
                job_source_type = job_source_type,
                final_reply_sent = final_reply_sent,
            )
        except Exception as e:
            self.logger.system(e, job_id)
            

    def _enter_degraded_loop(self) -> Never:
        '''Run essentials, where the priority is replying to user emails.'''  

        self.logger.system("running")
        self.friends_repo.reload_if_modified()

        
        while True:
            try:
                time.sleep(1)
                self._check_for_stop_flag()
                self._check_for_restart_flag()

                # process one personal inbox email in degraded mode
                paths = self.mail_backend_personal.list_inbox_mail_paths(max_items=1)
                if not paths:
                    continue
                
                inbox_path = paths[0]
                candidate = self.mail_backend_personal.parse_mail_file(inbox_path)                
                candidate = self.mail_backend_personal.move_to_processing(candidate)

                try: self.logger.ui(f"email from {candidate.email_address}", blank_line_before=True)
                except Exception: pass

                # silent delete non friends
                if not self.friends_repo.is_allowed_sender(candidate.email_address):
                    self.logger.ui("--> rejected (not in friends.xlsx)")
                    self.mail_backend_personal.delete_from_processing(candidate)
                    continue
                
                # check for email commands
                self._check_for_restart_command(candidate)
                self._check_for_stop_command(candidate)

                # reply, audit-log and delete for friends
                job_id = self._generate_job_id()
                final_reply_sent = self._try_send_out_of_service_reply(candidate, job_id)
                self._try_insert_recovery_audit_row(job_id, candidate, final_reply_sent, reason="SAFESTOP")
                
                try: self.logger.ui("--> rejected (safestop)")
                except Exception: pass
            
            except Exception as e:
                self.logger.system(e)

    def _restart_application(self) -> Never:
        ''' written by AI '''
        self.logger.system("restarting application in new visible terminal")

        try:
            self._set_ui_shutdown()
        except Exception:
            pass

        try:
            script_path = os.path.abspath(sys.argv[0])

            if platform.system() == "Windows":
                subprocess.Popen(
                    [sys.executable, script_path],
                    creationflags=subprocess.CREATE_NEW_CONSOLE # type: ignore
                )

            else:
                python_cmd = f'"{sys.executable}" "{script_path}"'

                terminal_candidates = [
                    ["gnome-terminal", "--", "bash", "-lc", f"{python_cmd}; exec bash"],
                    ["xfce4-terminal", "--hold", "-e", python_cmd],
                    ["konsole", "-e", "bash", "-lc", f"{python_cmd}; exec bash"],
                    ["xterm", "-hold", "-e", python_cmd],
                ]

                launched = False
                for cmd in terminal_candidates:
                    try:
                        subprocess.Popen(cmd)
                        launched = True
                        break
                    except FileNotFoundError:
                        continue

                if not launched:
                    raise RuntimeError("No supported terminal emulator found for restart")

        except Exception as e:
            self.logger.system(e)
            os._exit(1)

        time.sleep(1)
        os._exit(0)


# ============================================================
# UI
# ============================================================

class DashboardUI:
    """Tkinter dashboard for runtime status, logs, and operator visibility."""

    # colors
    BG = "#000000"
    TEXT = "#F5F5F5"
    MUTED = "#A0A0A0"
    GREEN = "#22C55E"
    GREEN_2 = "#16A34A"
    GREEN_3 = "#15803D"
    RED = "#DC2626"
    YELLOW = "#FACC15"
    SCROLL_TROUGH = "#0F172A"
    SCROLL_BG = "#1E293B"
    SCROLL_ACTIVE = "#475569"

    # fonts
    FONT_STATUS = ("Arial", 100, "bold")
    FONT_COUNTER = ("Segoe UI", 140, "bold")
    FONT_SMALL = ("Arial", 14, "bold")
    FONT_LOG = ("DejaVu Sans Mono", 20)
    FONT_RECORDING = ("Arial", 20, "bold")

    # sizes
    WINDOW_GEOMETRY = "1800x1000+0+0"
    ROOT_PADX = 50
    SCROLLBAR_WIDTH = 23

    RECORDING_WIDTH = 250
    RECORDING_HEIGHT = 110
    RECORDING_MARGIN_RIGHT = 30


    def __init__(self, shutdown_callback=None):
        self.shutdown_callback = shutdown_callback
        self._build_root(self.BG)
        self._build_header(self.BG, self.TEXT)
        self._build_body(self.BG, self.TEXT)
        self._build_footer(self.BG, self.TEXT)

        #self._debug_grid(self.root)


    def run(self) -> None:
        self.root.mainloop()


    def set_shutdown_callback(self, callback) -> None:
        self.shutdown_callback = callback


    def shutdown(self) -> None:
        if self._closing:
            return

        self._closing = True

        try:
            if self.shutdown_callback is not None:
                self.shutdown_callback()
        except Exception:
            pass

        self.root.destroy()


    def _debug_grid(self, widget):
        ''' highlights all grids with red '''
        for child in widget.winfo_children():
            try:
                child.configure(highlightbackground="red", highlightthickness=1)
            except Exception:
                pass
            self._debug_grid(child)


    def _build_root(self, bg_color):
        self.root = tk.Tk()
        #self.root.geometry(self.WINDOW_GEOMETRY)
        #self.root.resizable(False, False)

        if platform.system() == "Windows":
            self.root.state("zoomed")
        else:
            self.root.attributes("-zoomed", True)

        self.root.configure(bg=bg_color, padx=self.ROOT_PADX)
        self._closing = False
        self.root.protocol("WM_DELETE_WINDOW", self._on_close_attempt)

        self.root.title('RPA dashboard')
        self._create_recording_overlay()

        # layout using grid
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)


    def _build_header(self, bg_color, text_color):
        self.header = tk.Frame(self.root, bg=bg_color)

        self.header.grid(row=0, column=0, sticky="ew")
        self.header.grid_columnconfigure(2, weight=1)
        self.header.grid_rowconfigure(0, weight=1)

        # Header content
        self.rpa_text_label = tk.Label(
            self.header,
            text="RPA:",
            fg=text_color,
            bg=bg_color,
            font=self.FONT_STATUS,
        )  
        self.rpa_text_label.grid(row=0, column=0, padx=16, pady=16, sticky="w")

        self.rpa_status_label = tk.Label(
            self.header,
            text="",
            fg=self.RED,
            bg=bg_color,
            font=self.FONT_STATUS,
        )
        self.rpa_status_label.grid(row=0, column=1, padx=16, pady=16, sticky="w")

        self.status_dot = tk.Label(
            self.header,
            text="",
            fg=self.GREEN,
            bg=bg_color,
            font=("Arial", 50, "bold"),
        )
        self.status_dot.grid(row=0, column=2, sticky="w")

        # jobs done today (counter + label in same grid)
        self.jobs_counter_frame = tk.Frame(self.header, bg=bg_color)
        self.jobs_counter_frame.grid(row=0, column=3, sticky="ne", padx=40, pady=30)
        self.jobs_counter_frame.grid_rowconfigure(0, weight=1)
        self.jobs_counter_frame.grid_columnconfigure(0, weight=1)

        # normal view (jobs done today)
        self.jobs_normal_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_normal_view.grid(row=0, column=0, sticky="nsew")
        self.jobs_normal_view.grid_columnconfigure(0, weight=1)

        self.jobs_done_label = tk.Label(
            self.jobs_normal_view,
            text="0",
            fg=text_color,
            bg=bg_color,
            font=self.FONT_COUNTER,
            anchor="e",
            justify="right",
        )
        self.jobs_done_label.grid(row=0, column=0, sticky="e")

        self.jobs_counter_text = tk.Label(
            self.jobs_normal_view,
            text="jobs done today",
            fg=self.MUTED,
            bg=bg_color,
            font=self.FONT_SMALL,
            anchor="e",
        )
        self.jobs_counter_text.grid(row=1, column=0, sticky="e", pady=(0, 6))

        # safestop view (big X)
        self.jobs_error_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_error_view.grid(row=0, column=0, sticky="nsew")

        self.safestop_x_label = tk.Label(
            self.jobs_error_view,
            text="X",
            bg=self.RED,
            fg="#FFFFFF",
            font=self.FONT_COUNTER,
        )  # text="✖",
        self.safestop_x_label.pack(expand=True)

        # show normal view at startup
        self.jobs_normal_view.tkraise()

        # 'online'-status animation
        self._online_animation_after_id = None
        self._online_pulse_index = 0

        # 'working...'-status animation
        self._working_animation_after_id = None
        self._working_dots = 0


    def _build_body(self, bg_color, text_color):
        self.body = tk.Frame(self.root, bg=bg_color)
        self.body.grid(row=1, column=0, sticky="nsew")
        self.body.grid_rowconfigure(0, weight=1)
        self.body.grid_columnconfigure(0, weight=1)

        # body content
        log_and_scroll_container = tk.Frame(self.body, bg=bg_color)
        log_and_scroll_container.grid(row=0, column=0, sticky="nsew")
        log_and_scroll_container.grid_rowconfigure(0, weight=1)
        log_and_scroll_container.grid_columnconfigure(0, weight=1)

        # the right-hand side scrollbar
        scrollbar = tk.Scrollbar(
            log_and_scroll_container,
            width=self.SCROLLBAR_WIDTH,
            troughcolor=self.SCROLL_TROUGH,
            bg=self.SCROLL_BG,
            activebackground=self.SCROLL_ACTIVE,
            bd=0,
            highlightthickness=0,
            relief="flat",
        )
        scrollbar.grid(row=0, column=1, sticky="ns")

        # the 'console'-style log
        self.log_text = tk.Text(
            log_and_scroll_container,
            yscrollcommand=scrollbar.set,
            bg=bg_color,
            fg=text_color,
            insertbackground="black",
            font=self.FONT_LOG,
            wrap="none",
            state="disabled",
            bd=0,
            highlightthickness=0,
        )  # glow highlightbackground="#1F2937", highlightthickness=1
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.config(command=self.log_text.yview)


    def _build_footer(self, bg_color, text_color):
        self.footer = tk.Frame(self.root, bg=bg_color)
        self.footer.grid(row=2, column=0, sticky="nsew")
        self.footer.grid_rowconfigure(0, weight=1)
        self.footer.grid_columnconfigure(0, weight=1)

        # footer content
        self.last_activity_label = tk.Label(
            self.footer,
            text="last activity: xx:xx",
            fg=self.MUTED,
            bg=bg_color,
            font=self.FONT_SMALL,
            anchor="e",
        )
        self.last_activity_label.grid(row=0, column=1, padx=8, pady=16)


    def _apply_status_update(self, status: UIStatus | None = None):

        # stops any ongoing animations
        self._stop_online_animation()
        self._stop_working_animation()
        self.status_dot.config(text="")

        # changes text
        if status == "online":
            self.rpa_status_label.config(text="online", fg=self.GREEN)
            self.jobs_normal_view.tkraise()
            self.status_dot.config(text="●")
            self._start_online_animation()

        elif status == "no_network":
            self.rpa_status_label.config(text="no network", fg=self.RED)
            self.jobs_normal_view.tkraise()

        elif status == "working":
            self.rpa_status_label.config(text="working...", fg=self.YELLOW)
            self.jobs_normal_view.tkraise()
            self._start_working_animation()

        elif status == "safestop":
            self.rpa_status_label.config(text="safestop", fg=self.RED)
            self.jobs_error_view.tkraise()

        elif status == "out_of_office":
            self.rpa_status_label.config(text="out of office", fg=self.YELLOW)
            self.jobs_normal_view.tkraise()


    def _apply_jobs_done_today(self, n) -> None:
        self.jobs_done_label.config(text=str(n))


    def _create_recording_overlay(self) -> None:
        #written by AI
        self.recording_win = tk.Toplevel(self.root)
        self.recording_win.withdraw()                # hidden at start
        self.recording_win.overrideredirect(True)    # no title/border
        self.recording_win.configure(bg="black")

        try:
            self.recording_win.attributes("-topmost", True)
        except Exception:
            pass

        width = self.RECORDING_WIDTH
        height = self.RECORDING_HEIGHT
        x = self.root.winfo_screenwidth() - width - self.RECORDING_MARGIN_RIGHT
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

        frame = tk.Frame(
            self.recording_win,
            bg="black",
            highlightbackground="#444444",
            highlightthickness=1,
            bd=0,
        )
        frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(
            frame,
            width=44,
            height=44,
            bg="black",
            highlightthickness=0,
            bd=0,
        )
        canvas.place(x=18, y=33)
        canvas.create_oval(4, 4, 40, 40, fill=self.RED, outline=self.RED)

        label = tk.Label(
            frame,
            text="RECORDING",
            fg="#FFFFFF",
            bg="black",
            font=self.FONT_RECORDING,
            anchor="w",
        )
        label.place(x=75, y=33)


    def _show_recording_overlay(self) -> None:
        #written by AI
        try:
            width = self.RECORDING_WIDTH
            height = self.RECORDING_HEIGHT
            x = self.root.winfo_screenwidth() - width - self.RECORDING_MARGIN_RIGHT
            y = (self.root.winfo_screenheight() // 2) - (height // 2)
            self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

            self.recording_win.deiconify()
            self.recording_win.lift()

            try:
                self.recording_win.attributes("-topmost", True)
            except Exception:
                pass
        except Exception:
            pass


    def _hide_recording_overlay(self) -> None:
        # hides recording window
        try:
            self.recording_win.withdraw()
        except Exception:
            pass


    def _start_working_animation(self):
        if self._working_animation_after_id is None:
            self._animate_working()


    def _animate_working(self):
        #written by AI
        states = ["working", "working.", "working..", "working..."]
        self._working_dots = (self._working_dots + 1) % len(states)
        self.rpa_status_label.config(text=states[self._working_dots])
        self._working_animation_after_id = self.root.after(500, self._animate_working)


    def _stop_working_animation(self):
        if self._working_animation_after_id is not None:
            self.root.after_cancel(self._working_animation_after_id)
            self._working_animation_after_id = None
            self._working_dots = 0


    def _start_online_animation(self):
        if self._online_animation_after_id is None:
            self._online_pulse_index = 0
            self._animate_online()


    def _animate_online(self):
        # green pulse animation
        colors = [self.GREEN, self.GREEN_2, self.BG, self.GREEN_3, self.GREEN_2]
        color = colors[self._online_pulse_index]

        self.status_dot.config(fg=color)

        self._online_pulse_index = (self._online_pulse_index + 1) % len(colors)
        self._online_animation_after_id = self.root.after(1000, self._animate_online)


    def _stop_online_animation(self):
        if self._online_animation_after_id is not None:
            self.root.after_cancel(self._online_animation_after_id)
            self._online_animation_after_id = None


    def _append_ui_log(self, log_line: str, blank_line_before: bool = False) -> None:

        self.log_text.config(state="normal")  # open for edit
        now = datetime.datetime.now().strftime("%H:%M")

        if blank_line_before:
            self.log_text.insert("end", "\n")

        self.log_text.insert("end", f"[{now}] {log_line}\n")

        self.log_text.config(state="disabled")  # closing edit
        self.log_text.see("end")

    def _on_close_attempt(self):
        self.post_log_line("Use STOP-button in RPA tool next time! (or press 2 in rpa_tool_simulator.py)", blank_line_before=True)
        print("Use STOP-button in RPA tool next time! (or press 2 in rpa_tool_simulator.py)")
        self.post_shutdown(delay=2000)

    # all 'post_...' below are thread-safe wrappers
    def post_status_update(self, status: UIStatus) -> None:
        self.root.after(0, lambda: self._apply_status_update(status))

    def post_log_line(self, text: str, blank_line_before: bool = False) -> None:
        self.root.after(0, lambda: self._append_ui_log(text, blank_line_before))

    def post_show_recording_overlay(self) -> None:
        self.root.after(0, self._show_recording_overlay)

    def post_hide_recording_overlay(self) -> None:
        self.root.after(0, self._hide_recording_overlay)

    def post_jobs_done_today(self, n: int) -> None:
        self.root.after(0, lambda: self._apply_jobs_done_today(n))

    def post_shutdown(self, delay=0) -> None:
        self.root.after(delay, self.shutdown)


# ============================================================
# MAIN ENTRYPOINT
# ============================================================

class RobotRuntime:
    """Main orchestration runtime."""

    RPA_TOOL_CLAIM_TIMEOUT = 10 # Max wait time for RPA tool to claim workflow
    RPA_TOOL_EXECUTION_TIMEOUT = 600 # Max wait time for RPA tool to finish workflow
    POLL_INTERVAL = 1   # demo-friendly poll interval for runtime_loop() 
    QUERYFLOW_POLLINTERVAL = 1  # demo-friendly poll interval for queries (seconds)

    # eg. operating hours 05:00 to 23:00
    OPERATING_HOURS_START = datetime.time(5, 0)
    OPERATING_HOURS_END = datetime.time(23, 0)

    def __init__(self, ui):

        self.prev_ui_status = None
        self.next_queryflow_check_time = 0
        self.prev_state: HandoverState | None = None
        self.rpa_tool_claim_started_at: float | None = None
        self.rpa_tool_execution_started_at: float | None = None

        self.ui = ui
        self.logger = LoggerService(self.ui)

        self.handover_repo = HandoverRepository(self.logger)
        self.friends_repo = FriendsRepository()
        self.audit_repo = AuditRepository(self.logger)
        self.network_service = NetworkService(self.logger)
        self.recording_service = RecordingService(self.logger)
        
        self.mail_backend_personal = ExampleMailBackend(self.logger, "personal_inbox")
        self.mail_backend_shared = ExampleMailBackend(self.logger, "shared_inbox")
        self.erp_backend = ExampleErpBackend()

        self.job_handlers = {
            "ping": ExamplePingJobHandler(self.logger),
            "job1": ExampleJob1Handler(self.logger), 
            "job2": ExampleJob2Handler(self.logger), 
            "job3": ExampleJob3Handler(self.logger, self.erp_backend),
            }
        
        self.notification_service = UserNotificationService(self.mail_backend_personal, self.recording_service.RECORDINGS_DESTINATION_FOLDER, self.RPA_TOOL_EXECUTION_TIMEOUT)
        self.pre_handover_service = PreHandoverService(self.logger, self.handover_repo, self._update_ui_status, self.ui.post_show_recording_overlay, self._generate_job_id, self.recording_service, self.audit_repo, self.notification_service, self.mail_backend_personal, self.mail_backend_shared,)
        self.mail_flow = MailFlow(self.logger, self.friends_repo, self._is_within_operating_hours, self.network_service, self.job_handlers, self.pre_handover_service, self.mail_backend_personal, self.mail_backend_shared)
        self.query_flow = QueryFlow(self.logger, self.audit_repo, self.job_handlers, self.pre_handover_service, self._is_within_operating_hours, self.erp_backend)
        self.post_handover_service = PostHandoverService(self.logger, self.audit_repo, self.job_handlers, self.recording_service, self.ui.post_hide_recording_overlay, self.mail_backend_personal, self.mail_backend_shared, self.notification_service)
        self.safestop_controller = SafeStopController(self.logger, self.recording_service, self.ui.post_hide_recording_overlay, self.ui.post_status_update, self.ui.post_shutdown, self.mail_backend_personal, self.audit_repo, self._generate_job_id, self.friends_repo, self.notification_service, self._check_for_stop_flag,) 

        
    def _initialize_runtime(self,):
        handover_job: HandoverJob | None = None
        try:
            self.logger.system(f"RobotRuntime started, version={VERSION}, pid={os.getpid()}")

            # cleanup
            for fn in ["stop.flag", "restart.flag"]:
                try: os.remove(fn)
                except Exception: pass
       
            # write 'idle' to allow for manual start of main.py (not the intended way)
            self.handover_repo.write(HandoverJob(state="idle"))

            handover_job = self.handover_repo.read()
            if handover_job.state != "idle":
                raise PreHandoverCrash(
                    f"Expected handover.json to start in idle, got {handover_job.state}",
                    handover_job=handover_job,
                )        

            atexit.register(self.recording_service.stop) # during normal exit
    
            self.network_service.has_network_access()
            self.recording_service.stop() # stop any active recordings since last session
            self.recording_service.cleanup_aborted_recordings()
            self.friends_repo.reload_if_modified()
            self.audit_repo.ensure_db_exists()
            self._refresh_jobs_done_counter()
            
            self.safestop_controller.recover_pending_mail_jobs(from_initialize=True) 
            
    
        except PreHandoverCrash:
            raise

        except Exception as e:
            raise PreHandoverCrash(
                f"_initialize_runtime failed: {e}",
                handover_job=handover_job,
                cause=e,
            ) from e


    def runtime_loop(self) -> None:
        handover_job: HandoverJob | None = None

        try:
            self._initialize_runtime()
            
            while True:
                self._check_for_stop_flag()

                handover_job = self.handover_repo.read()
                self._handle_state_transition(handover_job)
                self._enforce_watchdog(handover_job)

                state = handover_job.state
                job_id = handover_job.job_id
                
                # dispatch
                if state == "idle":             # RobotRuntime owns the workflow
                    self._poll_job_intake()

                elif state == "job_queued":     # RPA Tool owns the workflow
                    pass

                elif state == "job_running":    # RPA Tool owns the workflow
                    pass

                elif state == "job_verifying":  # RobotRuntime owns the workflow
                    self._finalize_current_job(handover_job)

                elif state == "safestop":       # RobotRuntime owns the workflow
                    raise RpaToolCrash(
                        "crash signal received from RPA tool",
                        job_id=job_id,
                        handover_job=handover_job,
                    ) 

                time.sleep(self.POLL_INTERVAL)


        except RuntimeFault as fault:
            fault.traceback_text = traceback.format_exc()
            self.safestop_controller.run_degraded_mode(fault)

        except Exception as err:
            fault = RuntimeFault(
                message=str(err),
                job_id=handover_job.job_id if handover_job else None,
                handover_job=handover_job,
                cause=err,
                traceback_text=traceback.format_exc(),
            )
            self.safestop_controller.run_degraded_mode(fault)


    def _refresh_jobs_done_counter(self, job_id=None):
        try:
            count = self.audit_repo.count_done_jobs_today()
            self.ui.post_jobs_done_today(count)
        except Exception as err:
            self.logger.system(err, job_id)


    def _handle_state_transition(self, handover_job: HandoverJob) -> None:
        job_id = handover_job.job_id
        state = handover_job.state

        if state != self.prev_state:
            transition_message=f"state transition detected by CPU-poll: {self.prev_state} -> {state}"

            if not self.handover_repo.is_valid_transition(self.prev_state, state):
                raise RuntimeError(f"invalid {transition_message}")

            self._update_ui_status(state)
            self.logger.system(transition_message, job_id)

            if state == "job_running":
                self.audit_repo.update_job(job_id=job_id, job_status="RUNNING")

            # note handover time or last RPA tool state transition
            if state == "job_queued":
                self.rpa_tool_claim_started_at = time.time()
                self.rpa_tool_execution_started_at = None

            elif state == "job_running":
                self.rpa_tool_claim_started_at = None
                self.rpa_tool_execution_started_at = time.time()

            else:
                self.rpa_tool_claim_started_at = None
                self.rpa_tool_execution_started_at = None
        
        self.prev_state = state


    def _enforce_watchdog(self, handover_job):
        state = handover_job.state
        now = time.time()

        if state == "job_queued":
            if self.rpa_tool_claim_started_at is None:
                return

            if now - self.rpa_tool_claim_started_at > self.RPA_TOOL_CLAIM_TIMEOUT:
                raise PreHandoverCrash(
                    f"Your request was correct, but the robot was somehow unable to start ({self.RPA_TOOL_CLAIM_TIMEOUT} seconds limit).",
                    job_id=handover_job.job_id,
                    handover_job=handover_job,
                )

        elif state == "job_running":
            if self.rpa_tool_execution_started_at is None:
                return
        
            if now - self.rpa_tool_execution_started_at > self.RPA_TOOL_EXECUTION_TIMEOUT:
                raise RpaToolCrash(
                    f"Unable to finish the job within {self.RPA_TOOL_EXECUTION_TIMEOUT} seconds timeout",
                    job_id=handover_job.job_id,
                    handover_job=handover_job,
                )


    def _update_ui_status(self, state=None, forced_status=None) -> None:
               
        if forced_status is not None:
            if forced_status not in get_args(UIStatus):
                raise ValueError(f"unknown forced_status: {forced_status}")
            ui_status: UIStatus = forced_status

        else:
            if state is not None and state not in get_args(HandoverState):
                raise ValueError(f"unknown state: {state}")

            if state == "safestop":
                ui_status = "safestop"

            elif state in ("job_queued", "job_running", "job_verifying"):
                ui_status = "working"

            elif self.network_service.network_state is False:
                ui_status = "no_network"

            elif not self._is_within_operating_hours():
                ui_status = "out_of_office"

            else:
                ui_status = "online"

        if self.prev_ui_status != ui_status:
            self.ui.post_status_update(ui_status)
            self.prev_ui_status = ui_status


    def _poll_job_intake(self) -> bool:
        ''' job intake logic '''
        try:
            
            # 1. Mail first (priority)
            if self.mail_flow.poll_once():                
                return True
            
            # 2. Query (or other scheduled) jobs
            now = time.time()
            if now < self.next_queryflow_check_time:
                return False

            if self.query_flow.poll_once():
                return True

            # prolong interval if no new match
            self.next_queryflow_check_time = now + self.QUERYFLOW_POLLINTERVAL 
            return False


        except PreHandoverCrash:
            raise
        except Exception as e:
            raise PreHandoverCrash(str(e), cause=e) from e

        
    def _generate_job_id(self) -> int:
        ''' unique id for all jobs. This works under single-runtime-single-machine assumption'''

        job_id = int(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))

        last_job_id = self.audit_repo.get_latest_job_id()
        job_id = max(job_id, last_job_id + 1)

        self.logger.system(f"assigned job_id", job_id)
        return job_id

    
    def _is_within_operating_hours(self) -> bool:
        now = datetime.datetime.now().time()
        return self.OPERATING_HOURS_START <= now <= self.OPERATING_HOURS_END 
        

    def _finalize_current_job(self, handover_job: HandoverJob) -> None:
        job_id = handover_job.job_id

        self.logger.system(f"finalizing {handover_job.job_type} with payload {handover_job.rpatool_payload}", job_id) # only store safe data in log
        
        self.post_handover_service.finalize_handover_job(handover_job)
        self._refresh_jobs_done_counter(job_id)
        
        try:
            self.handover_repo.write(
                HandoverJob(
                    state="idle"
                    ))
            
        except Exception as err:
            raise PostHandoverCrash(
                "unable to reset handover.json",
                job_id=job_id,
                handover_job=handover_job,
                cause=err,
                )


    def _check_for_stop_flag(self):
        ''' to stop main.py on operator manual stop on RPA tool '''

        stopflag = "stop.flag"
 
        if os.path.isfile(stopflag):
            try: os.remove(stopflag)
            except Exception: pass

            self.logger.system(f"found {stopflag}")
            
            try: self.ui.post_shutdown() #request soft-exit
            except Exception: os._exit(1)
            
            time.sleep(3)
            os._exit(0)  #kill if still alive after 3 sec 


    def request_shutdown(self) -> None:
        ''' stop recording at controlled UI shutdown'''
        try:
            self.recording_service.stop()
        except Exception:
            pass


def main() -> None:
    ''' run Dashboard UI in main thread and 'the rest' async '''
    ui = DashboardUI()
    robot_runtime = RobotRuntime(ui)

    ui.set_shutdown_callback(robot_runtime.request_shutdown)

    threading.Thread(target=robot_runtime.runtime_loop, daemon=True).start() # 'the rest'

    ui.run()


if __name__ == "__main__":
    main()